from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from tkinter import font
from os import remove as PATHREMOVE
from shutil import rmtree as REMOVEDIR
from os import path
from os import mkdir as NOVODIRETORIO
from shutil import copy as COPYFILE
from openpyxl import *
from datetime import date
from datetime import datetime
from datetime import time
from tkcalendar import Calendar
from math import ceil
from PIL import Image, ImageTk


class Stack:
    def __init__(self):

        self.__Stack = []

    def Adicionar(self, valor):

        self.__Stack = [valor] + self.__Stack

    def Remover(self):

        if self.Tamanho() > 0:
            self.__Stack = self.__Stack[1:]
        else:
            return False

    def Primeiro(self):

        if self.Tamanho() > 0:
            return self.__Stack[0]
        else:
            return None

    def Segundo(self):

        if self.Tamanho() > 1:
            return self.__Stack[1]
        else:
            return None

    def Ultimo(self):

        if self.Tamanho() > 0:
            return self.__Stack[-1]
        else:
            return None

    def Tamanho(self):

        return len(self.__Stack)

    def g(self):
        return self.__Stack


class Settings:

    # Program Vars
    __Languages = ("PT", "ING")
    __Lyts = ("Left", "Right", "Up", "Down", "Custum")
    __Numero_de_Aberturas = 0
    __Versao = "0.0.0000001"
    __LastSession = None
    ## __Defaults_Layouts = {"Custum":{},"Left":{"X":0,"Y":0,"W":0,"H":0},"Right":{},"Up":{},"Down":{}}
    # Program Defaults
    __Defaults = {
        "Font_Menu": {"Font": "", "T": 10, "FG": None, "BG": None},
        "Font_Text1": {"Font": "", "T": 10, "FG": "White", "BG": None},
        "Font_Text2": {"Font": "", "T": 20, "FG": "pink", "BG": None},
        "Font_Text3": {"Font": "", "T": 10, "FG": "pink", "BG": None},
        "Font_Buttons": {"Font": "", "T": 10, "FG": None, "BG": None},
        "Tab_Font_1": {"Font": "", "T": 10, "FG": None, "BG": None},
        "Tab_Font_2": {"Font": "Courier", "T": 12, "FG": None, "BG": None},
        "Tab_Font_3": {"Font": "", "T": 10, "FG": "pink", "BG": None},
        "Buttons_congig": {"FG": None, "BG": "#f0f0f0", "Selected": "#4dd2ff"},
        "Lyt_Clrs_Tab": {"BG": "", "FG": "", "Fundo": "White"},
        "Lyt_Clrs_Lista": {},
        "Lyt_Clrs_Tebela": {
            "Top_BG": "Green",
            "Top_L": "",
            "Grelha1": "",
            "Grelha2": "",
            "FotoSpace": "",
            "BG": None,
        },
        "Lyt_Clrs_Bar": {},
        "Lyt_Clrs_Buttons": {"Sel": "", "BR": "", "IN": ""},
        "Path": "",
        "Path_Images": "",
        "Path_Winfo_Images": ("", "", "", ""),
        "Winfo_Images": {"Font": "", "T": 10, "FG": None, "BG": None},
        "Idioma": "PT",
        "Layout": "Left",
    }

    __User_Vars = __Defaults
    ##    __User_Layout = __Defaults_Layouts["Left"]

    @classmethod
    def LastSession(cls):
        return cls.__LastSession

    @classmethod
    def Versao(cls):
        return cls.__Versao

    @classmethod
    def GetOpens(cls):
        return cls.__Numero_de_Aberturas

    @classmethod
    def Save(cls):

        pass

    @classmethod
    def Load(cls):
        pass

    @classmethod
    def Get(cls, var):
        """
        Devolve o valor pedido

        Fontes de letra ["Font_Menu", "Font_Txt1", "Font_Txt2", "Font_Txt3"]
        Cores           ["Lyt_Clrs_Tab", "Lyt_Clrs_Lista", "Lyt_Clrs_Tebela", "Lyt_Clrs_Bar", "Lyt_Clrs_Buttons"]
        Outros          ["Path", "Path_Images", "Idioma", "Layout"]
        """

        if var in cls.__Defaults:
            return cls.__User_Vars[var]
        else:
            raise ValueError("Argumento inválido: não existe a variavel:", var)

    @classmethod
    def Set(cls, key, valor):
        """
        Modifica destrutivamente a variavel em questão e verifica o valor ,retorna True se foi feita alteração,
        """
        if cls.__Compardor(53, key, valor):
            cls.__User_Vars[key] = valor
        else:
            # Notificação sem alteração
            pass

    @classmethod
    def __Compardor(cls, linha, key, valor):

        if key in cls.__Defaults:

            D_valor = cls.__Defaults[key]

            if isinstance(D_valor, int):
                return cls.__VerifyInt(linha, valor)
            elif isinstance(D_valor, str):

                if key == "Idioma" and valor in cls.__Languages:
                    return True
                elif key == "Layout" and valor in cls.__Lyts:
                    return True
                elif key in ("Path", "Path_Images") and path.isdir(valor):
                    return True
                elif cls.__VerifyHex(linha, valor):
                    return True
                else:
                    # valor enrrado linha
                    return False

            elif isinstance(D_valor, dict):
                return cls.__VerifyDict(linha, D_valor, valor)
            else:
                # valor enrradolinha
                return False

        else:
            # Chave Inválida
            return False

    @classmethod
    def __VerifyDict(cls, linha, D_dic, dic):

        c = 0
        for key in D_dic:

            if key == "Font":

                if dic[key] in font.families():
                    c += 1

            elif isinstance(D_dic[key], str) and isinstance(dic[key], str):

                if cls.__VerifyHex(linha, dic[key]):
                    c += 1

            elif isinstance(D_dic[key], int) and isinstance(dic[key], int):

                if dic[key] > 0:
                    c += 1

            else:
                c = c

        if c == len(D_dic):
            return True
        else:
            # runtime error linha string: não são semelhantes
            return False

    @classmethod
    def __VerifyHex(cls, linha, string):

        if len(string) == 6:

            for carater in string.upper():

                if carater not in (
                    "0",
                    "1",
                    "2",
                    "3",
                    "4",
                    "5",
                    "6",
                    "7",
                    "8",
                    "9",
                    "A",
                    "B",
                    "C",
                    "D",
                    "E",
                    "F",
                    "#",
                ):
                    return False

            return True

        else:
            # runtime error linha string: string não é Hex
            return False

    @classmethod
    def __VerifyInt(cls, linha, inte):

        if isinstance(inte, int):
            return inte > 0
        else:
            # runtime error linha interio: inte não é interio positivo
            return False


class MyWidgets:
    class Zoom_Bar(Canvas):
        def __init__(self, master, minimo, maximo, step):

            super().__init__(master=master, background="white", highlightthickness=0)

            self.__min = minimo
            self.__max = maximo
            self.__step = step

            # menos
            self.create_oval((2, 2, 17, 17), fill="red", tag=("Bt_menos"))

            # mais
            self.create_oval((123, 2, 138, 17), fill="red", tag=("Bt_mais"))

            # barra
            self.create_rectangle((20, 8, 120, 11), fill="black")

            # "cursor"
            self.create_rectangle(
                (20 + 57, 2, 20 + 57 + 3, 17), fill="gray", tag=("Bt_slide")
            )

    class Progress_Bar:  # não atualiza posição

        __Eventos = []

        @classmethod
        def Bind(cls, func):

            cls.__Eventos += [func]

        @classmethod
        def UnBind(cls, func):

            for i in range(0, len(cls.__Eventos)):

                if cls.__Eventos[i] == func:

                    del cls.__Eventos[i]

        def __RunEvent(self, valor, m):

            for sub in self.__Eventos:

                sub(valor, m)

        def __init__(
            self,
            Master,
            posx,
            posy,
            w=100,
            h=12,
            steps=1,
            brwidth=2,
            brcolor="#d9d9d9",
            barcolor="#7bb900",
            Text="teste",
            wl=200,
            BG="red",
        ):

            self.__Master = Master
            self.__brwidth = brwidth
            self.__largura = w - self.__brwidth
            self.__step = (self.__largura) / steps
            self.__steps = steps
            self.__heigh = h - self.__brwidth

            self.__Progress = 0

            self.__Label = Label(Master, text=Text, wraplength=wl, bg=BG)
            self.__Label.place(x=posx + self.__largura + self.__brwidth, y=posy - 4)

            self.__Canvas = Canvas(
                Master, width=w, height=h, background=brcolor, highlightthickness=0
            )
            self.__Canvas.create_rectangle(
                self.__brwidth,
                self.__brwidth,
                self.__brwidth,
                self.__heigh,
                fill=barcolor,
                outline="",
                tag="rectangle",
            )
            self.__Canvas.place(x=posx, y=posy)

            self.__RunEvent(True, self.__Master)

        def Progredir(self, peso=1, text="Estado"):

            self.__Progress += peso

            if not isinstance(peso, int):
                peso = 0

            largura = self.__brwidth + (self.__Progress) * self.__step

            if largura > self.__largura:
                self.__Canvas.coords(
                    "rectangle",
                    self.__brwidth,
                    self.__brwidth,
                    self.__largura,
                    self.__heigh,
                )

            else:
                self.__Canvas.coords(
                    "rectangle", self.__brwidth, self.__brwidth, largura, self.__heigh
                )

        def Reset(self):

            self.__Progress = 0
            self.__Canvas.coords(
                "rectangle",
                self.__brwidth,
                self.__brwidth,
                self.__brwidth + (self.__Progress) * self.__step,
                self.__heigh,
            )

        def Replace(self, posx, posy):

            self.__Canvas.place_forget()
            self.__Canvas.place(x=posx, y=posy)

        def Config_Desin(
            self, brwidth=2, brcolor="#d9d9d9", barcolor="#7bb900", text=""
        ):
            # terminar

            self.__brwidth = brwidth
            self.__largura = w - self.__brwidth
            self.__step = (self.__largura) / steps
            self.__steps = steps
            self.__heigh = h - self.__brwidth

        def Get_Progress(self):

            return round(self.__Progress / self.__steps * 100)

        def Set_Steps(self, steps):

            self.__step = self.__largura / steps

        def Destroy(self):

            self.__RunEvent(False, self.__Master)

            self.__Canvas.destroy()

    class Tab(Frame):

        Path = None

        def __init__(self, master, nome, path=None):

            super().__init__(master=master, relief=FLAT, bg="white", borderwidth=0)

            self.path = path
            # colunas
            self.columnconfigure(0, minsize=470, weight=1, uniform="A")
            self.columnconfigure(1, minsize=30, weight=0)
            self.columnconfigure(2, minsize=300, weight=1, uniform="A")
            # linhas
            self.rowconfigure(1, minsize=30, weight=0)
            self.rowconfigure(0, weight=1)

            # items
            self.tabela = MyWidgets.Tabela(self, nome)
            self.frm_InformaçõesTabela = MyWidgets.Frame_InformaçõesTabela(
                self, self.tabela
            )
            self.tabela.bind("<<Selecionado>>", self.__ON_Selecionado)
            self.tabela.bind("<<Desselecionado>>", self.__ON_Desselessionado)
            self.tabela.bind("<<ENDANIM>>", self.frm_InformaçõesTabela.Atualizar)
            master.bind("<<Tab_Save>>", self.__Save)

            # Frame Adicionar
            self.frm_Adicionar = MyWidgets.Frame_Adicionar(self, self.tabela)
            self.bind("<<click_adicionar>>", self.frm_InformaçõesTabela.Atualizar)

            # Frame Editar
            self.frm_Editar = MyWidgets.Frame_Editar(self, self.tabela)

            # Frame Ficha Técnica
            self.Ficha_tecnica = None

            # MenuVertcalBar
            self.menuBar = MyWidgets.MenuBarVertical(self)
            self.menuBar.bind("<<Button_Choiche>>", self.__action_click)

            # Searche Bar
            self.searchBar = MyWidgets.BarraDePrucura(self)

            # grids
            self.tabela.grid(row=0, column=0, sticky=NSEW)
            self.frm_InformaçõesTabela.grid(row=0, rowspan=2, column=2, sticky=NSEW)
            self.menuBar.grid(row=0, column=1, sticky=NSEW)
            self.searchBar.grid(row=1, column=0, columnspan=2, sticky=NSEW)

        def __Save(self, event):

            if event.x == 2:

                try:

                    if (
                        MyWidgets.Tab.Path != None
                        and MyWidgets.Tab.Path[MyWidgets.Tab.Path.find("?") + 1 :]
                        == self.tabela.Get_Nome()
                    ):

                        path = MyWidgets.Tab.Path[: MyWidgets.Tab.Path.find("?")]
                        Element_lst = self.tabela.Get_Elementos()

                        Wbook = Workbook()
                        Wsheet1 = Wbook.active
                        Wsheet1.title = "Livros"
                        Wsheet1["A1"] = "Versão 0.0.1"
                        Wsheet1["A2"] = "Id"
                        Wsheet1["B2"] = "Nome"
                        Wsheet1["C2"] = "Autor"
                        Wsheet1["D2"] = "Editora"
                        Wsheet1["E2"] = "Género"
                        Wsheet1["F2"] = "Edição"
                        Wsheet1["G2"] = "Data Recebimento"
                        Wsheet1["H2"] = "Data lançamaento"
                        Wsheet1["I2"] = "Condição"
                        Wsheet1["J2"] = "Outros"
                        Wsheet1["K2"] = "Número"

                        Wsheet2 = Wbook.create_sheet("Image Paths")
                        Wsheet2["A1"] = "Versão 0.0.1"
                        Wsheet2["A2"] = "Id's"
                        Wsheet2["B2"] = "Diretórios de imagens"

                        Wsheet3 = Wbook.create_sheet("Tabela info")
                        Wsheet3["A1"] = self.tabela.Get_Nome()

                        # historial
                        # emprestimos

                        c = 2
                        for Element_idd in Element_lst:
                            c += 1
                            elemento = self.tabela.Get_Elementos(
                                Element_idd, "Elemento"
                            )
                            livro = elemento.Get_livro()

                            numb = str(c)
                            Wsheet1["A" + numb] = str(Element_idd)
                            Wsheet1["B" + numb] = livro.Get_Title()
                            Wsheet1["C" + numb] = livro.Get_Autor()
                            Wsheet1["D" + numb] = livro.Get_Editora()
                            Wsheet1["E" + numb] = livro.Get_Genero()
                            Wsheet1["F" + numb] = livro.Get_Edition()
                            Wsheet1["G" + numb] = livro.Get_DataIn()
                            Wsheet1["H" + numb] = livro.Get_DataOut()
                            Wsheet1["I" + numb] = livro.Get_Condition()
                            Wsheet1["J" + numb] = livro.Get_Outros()
                            Wsheet1["k" + numb] = livro.Get_Numero()

                            Wsheet2["A" + numb] = str(Element_idd)

                            if (
                                livro.Get_ImagemPath() == None
                                or livro.Get_ImagemPath() == "n/d"
                            ):
                                Wsheet2["B" + numb] = "n/d"
                            else:
                                Wsheet2["B" + numb] = livro.Get_ImagemPath()
                                COPYFILE(
                                    livro.Get_ImagemPath(), path + "/Livros.imagens"
                                )
                                PATHREMOVE(livro.Get_ImagemPath())

                        path = path + "/" + path.split("/")[-1] + ".xlsx"
                        self.path = path
                        Wbook.save(path)
                        Wbook.close()

                except:

                    raise messagebox.showerror(
                        "Gestor de Livros", "Não foi possivel gravar."
                    )

            elif event.x == 1:

                if self.path != None:

                    name = self.path.split("/")[-2]
                    path2 = self.path[: -1 - len(self.path.split("/")[-1])]

                    REMOVEDIR(path2 + "/Livros.imagens")
                    NOVODIRETORIO(path2 + "/Livros.imagens")
                    PATHREMOVE(self.path)

                    Element_lst = self.tabela.Get_Elementos()

                    Wbook = Workbook()
                    Wsheet1 = Wbook.active
                    Wsheet1.title = "Livros"
                    Wsheet1["A1"] = "Versão 0.0.1"
                    Wsheet1["A2"] = "Id"
                    Wsheet1["B2"] = "Nome"
                    Wsheet1["C2"] = "Autor"
                    Wsheet1["D2"] = "Editora"
                    Wsheet1["E2"] = "Género"
                    Wsheet1["F2"] = "Edição"
                    Wsheet1["G2"] = "Data Recebimento"
                    Wsheet1["H2"] = "Data lançamaento"
                    Wsheet1["I2"] = "Condição"
                    Wsheet1["J2"] = "Outros"
                    Wsheet1["K2"] = "Número"

                    Wsheet2 = Wbook.create_sheet("Image Paths")
                    Wsheet2["A1"] = "Versão 0.0.1"
                    Wsheet2["A2"] = "Id's"
                    Wsheet2["B2"] = "Diretórios de imagens"

                    Wsheet3 = Wbook.create_sheet("Tabela info")
                    Wsheet3["A1"] = self.tabela.Get_Nome()

                    # historial
                    # emprestimos

                    c = 2
                    for Element_idd in Element_lst:
                        c += 1
                        elemento = self.tabela.Get_Elementos(Element_idd, "Elemento")
                        livro = elemento.Get_livro()

                        numb = str(c)
                        Wsheet1["A" + numb] = str(Element_idd)
                        Wsheet1["B" + numb] = livro.Get_Title()
                        Wsheet1["C" + numb] = livro.Get_Autor()
                        Wsheet1["D" + numb] = livro.Get_Editora()
                        Wsheet1["E" + numb] = livro.Get_Genero()
                        Wsheet1["F" + numb] = livro.Get_Edition()
                        Wsheet1["G" + numb] = livro.Get_DataIn()
                        Wsheet1["H" + numb] = livro.Get_DataOut()
                        Wsheet1["I" + numb] = livro.Get_Condition()
                        Wsheet1["J" + numb] = livro.Get_Outros()
                        Wsheet1["k" + numb] = livro.Get_Numero()

                        Wsheet2["A" + numb] = str(Element_idd)

                        if (
                            livro.Get_ImagemPath() == None
                            or livro.Get_ImagemPath() == "n/d"
                        ):
                            Wsheet2["B" + numb] = "n/d"
                        else:
                            Wsheet2["B" + numb] = livro.Get_ImagemPath()
                            COPYFILE(livro.Get_ImagemPath(), path2 + "/Livros.imagens")
                            PATHREMOVE(livro.Get_ImagemPath())

                    Wbook.save(self.path)
                    Wbook.close()

                else:
                    messagebox.showerror(
                        "Gestor de Livros",
                        "Nenhum diretorio disponivel -> Guardar em .",
                    )

        def __ON_Selecionado(self, event):

            self.menuBar.B2.configure(state="normal")
            livro = self.tabela.Get_Elementos(event.x, "Elemento").Get_livro()
            self.Ficha_tecnica = MyWidgets.Frame_FichaTecnica(self, self.tabela, livro)
            self.Ficha_tecnica.grid(row=0, rowspan=2, column=2, sticky=NSEW)

        def __ON_Desselessionado(self, event):

            self.menuBar.B2.configure(state="disable")
            if self.Ficha_tecnica != None:
                self.Ficha_tecnica.destroy()
                self.Ficha_tecnica = None

        def __action_click(self, event):

            if event.x == 1:

                self.frm_Adicionar.destroy()
                self.frm_Adicionar = MyWidgets.Frame_Adicionar(self, self.tabela)
                self.frm_Adicionar.grid(row=0, rowspan=2, column=2, sticky=NSEW)

            elif event.x == 2:

                if self.tabela.Get_Selected() != None:

                    self.tabela.Remover_Elemento(self.tabela.Get_Selected())
                    self.Ficha_tecnica.destroy()
                    self.frm_InformaçõesTabela.Atualizar(None)

            elif event.x == 3:

                self.frm_Editar.destroy()
                self.frm_Editar = MyWidgets.Frame_Editar(self, self.tabela)
                self.frm_Editar.grid(row=0, rowspan=2, column=2, sticky=NSEW)

            elif event.x == 4:
                pass

            #  if self.tabela.Get_Selected() != None:

            # livro = self.tabela.Get_Elementos(self.tabela.Get_Selected(),"Elemento").Get_livro()

            # self.Ficha_tecnica = MyWidgets.Frame_FichaTecnica(self,self.tabela,livro)
            # self.Ficha_tecnica.grid(row=0,rowspan=2,column=2,sticky=NSEW)

    class Tabela(Frame):
        def __init__(self, Master, nome, x=None, y=None):

            self.__Master = Master
            self.__nome = nome
            self.__Master.update_idletasks()

            self.__x = 0
            self.__y = y
            # self.__relwidth = relwidth
            # self.__relheight = relheight

            super().__init__(
                Master, bg=Settings.Get("Lyt_Clrs_Tebela")["Top_BG"], borderwidth=0
            )

            self.update_idletasks()
            self.__w = self.winfo_width()
            self.__h = self.winfo_height()

            self.update_idletasks()

            # self.place_configure(x=self.__x,y=self.__y,relwidth=self.__relwidth,relheight=self.__relheight)
            # Colunas
            self.columnconfigure(0, weight=0, pad=0)
            self.columnconfigure(1, weight=0, minsize=20, pad=0)
            self.columnconfigure(2, weight=0, minsize=100, pad=0)
            self.columnconfigure(3, weight=1, minsize=100, pad=0)
            # linhas
            self.rowconfigure(0, weight=0, minsize=30, pad=0)
            self.rowconfigure(1, weight=1, pad=0)

            # ScrollFrame e scrollbar
            self.__Canvas = Canvas(
                self,
                width=self.__w - 20,
                height=self.__h,
                highlightthickness=0,
                bg="pink",
            )
            self.__Canvas.configure(scrollregion=self.__Canvas.bbox("all"))
            self.__Canvas.grid(row=1, column=1, columnspan=3, sticky="NSEW")
            self.__ScrollBar = Scrollbar(
                self,
                orient="vertical",
            )
            self.__ScrollBar.grid(row=0, rowspan=2, column=0, sticky="NSEW")
            self.__Canvas.configure(yscrollcommand=self.__ScrollBar.set)
            self.__ScrollBar.configure(command=self.ScrollbarMove)

            self.__Canvas.bind("<Configure>", self.On_Resize)
            # self.__Canvas.bindtags("<MouseWheel>", self.On_MouseWheel )

            # Desins
            self.__lb_fantasma = Label(
                self, bg=Settings.Get("Lyt_Clrs_Tebela")["Top_BG"]
            )
            self.__lb_fantasma.grid(row=0, column=1, sticky="NSEW")

            self.__lb_Imagem = Label(
                self,
                text="Imagem",
                font=(
                    Settings.Get("Tab_Font_1")["Font"],
                    Settings.Get("Tab_Font_1")["T"],
                ),
                bg=Settings.Get("Lyt_Clrs_Tebela")["Top_BG"],
                fg=Settings.Get("Tab_Font_1")["FG"],
                justify=CENTER,
            )
            self.__lb_Imagem.grid(row=0, column=2, sticky="NSEW")

            self.__lb_MaisInfo = Label(
                self,
                text="Informações",
                font=(
                    Settings.Get("Tab_Font_1")["Font"],
                    Settings.Get("Tab_Font_1")["T"],
                ),
                bg=Settings.Get("Lyt_Clrs_Tebela")["Top_BG"],
                fg=Settings.Get("Tab_Font_1")["FG"],
                justify=CENTER,
            )
            self.__lb_MaisInfo.grid(row=0, column=3, sticky="NSEW")

            # Tabela vars
            self.__lst_OrdemDosEementos = []  # lista com os ids das canvas_window
            self.__lst_VisiveisAnterior = []
            self.__lst_Visiveis = []
            self.__lst_NotVisiveisAntes = []
            self.__lst_NotVisiveisDepois = []
            self.__dic_id_pos = {}  # dicionario de ids das canvas_window para livros
            self.__Last_id = 0
            self.__focuspos = 0
            self.__Element_Height = 140
            self.__MultiSelect = False
            self.__selected = None
            self.__animando = False
            self.__selecteds = []

            self.__ColorClara = "#AEAEAE"
            self.__ColorEscura = "#CDCDCD"

            self.__Ordenados = False

            # bin
            self.__AnimationQueue = {}

        def Get_Nome(self):
            return self.__nome

        def Get_Elemento_Height(self):
            return self.__Element_Height

        def Get_NumeroElementos(self):
            return len(self.__lst_OrdemDosEementos)

        def Get_Elementos(self, idd=None, optn=None):
            """
            self.__dic_id_pos[str(self.__Last_id)] = {"Elemento":elemento,"Windowid":idw,"Posid":self.Get_NumeroElementos()-1}
            optn = (Elemento,WWindowid,Posid)
            """
            if optn in ("Elemento", "Windowid", "Posid") and idd != None:
                return self.__dic_id_pos[str(idd)][optn]
            else:
                return self.__lst_OrdemDosEementos

        #   def Get_Selecionado(self):
        #      """
        #     retorna o idd do elemento da tabela selecionado, None se nenhum selecionado
        #    """
        #   return self.__selected

        def Get_Canvas(self):
            return self.__Canvas

        def Get_MultiSelect(self):
            return self.__MultiSelect

        def Get_Selected(self):

            """
            retorna o idd ou lista de idd's do elemento da tabela selecionado, None se nenhum selecionado
            """

            if self.__MultiSelect:
                return self.__selecteds
            else:
                return self.__selected

        def Desselecionar(self, idd):

            if str(idd) in self.__dic_id_pos:

                if self.__MultiSelect:
                    self.__selecteds.remove(idd)
                else:
                    self.__selected = None
                    self.__dic_id_pos[str(idd)]["Elemento"].Click_Reset()

                self.event_generate("<<Desselecionado>>", x=idd)

        def Set_Selected(self, idd):
            if self.__MultiSelect:

                if str(idd) in self.__dic_id_pos:
                    self.__selecteds += [
                        idd,
                    ]
                    self.event_generate("<<Selecionado>>", x=idd)
                elif idd == None:
                    self.__selecteds = []

            else:
                if str(idd) in self.__dic_id_pos:
                    self.__selected = idd
                    self.event_generate("<<Selecionado>>", x=idd)
                elif idd == None:
                    self.__selected = None

        def Atualizar_listas(self):
            """
            Atualiza a lista de ojetos potêncialmente visiveis para o utilizador depende da posição do primeiro e da ordem dos elementos
            # Quando não são eliminados os elementos antes de movidos lista de visisiveis fica incorreta pois de depende da ordem !!!
            """
            # update para não haver numeros errados
            self.__Canvas.update_idletasks()

            # calcular quantos se pode mostar
            q = ceil(self.__Canvas.winfo_height() / self.__Element_Height) + 2

            # calcula a posição min que é visivel
            yview_inicio = ceil(
                (self.__Canvas.yview()[0] * self.__focuspos) / self.__Element_Height - 1
            )

            # Atualizar aa anteriores:
            self.__lst_VisiveisAnterior = self.__lst_Visiveis

            if yview_inicio < 0:
                yview_inicio = 0

            if q >= len(self.__lst_OrdemDosEementos):
                self.__lst_Visiveis = self.__lst_OrdemDosEementos
                self.__lst_NotVisiveisAntes = []
                self.__lst_NotVisiveisDepois = []
            else:
                self.__lst_Visiveis = self.__lst_OrdemDosEementos[
                    yview_inicio : yview_inicio + q
                ]
                self.__lst_NotVisiveisAntes = self.__lst_OrdemDosEementos[:yview_inicio]
                self.__lst_NotVisiveisDepois = self.__lst_OrdemDosEementos[
                    yview_inicio + q :
                ]

            # print(q,len(self.__lst_Visiveis),self.__lst_Visiveis)

        def Atualizar_visiveis(self):
            """
            Atualiza as listas e torna visiveis os supostamente visiveis pelo utilizador
            """
            self.Atualizar_listas()

            for idd in self.__lst_Visiveis:
                if self.__dic_id_pos[str(idd)]["Windowid"] == None:
                    self.__dic_id_pos[str(idd)][
                        "Windowid"
                    ] = self.__Canvas.create_window(
                        0,
                        self.__dic_id_pos[str(idd)]["Elemento"].Get_Geometria()[1],
                        window=self.__dic_id_pos[str(idd)]["Elemento"],
                        anchor=NW,
                    )
                    print(
                        idd,
                        ":Criados:",
                        self.__dic_id_pos[str(idd)]["Elemento"].Get_Geometria()[1],
                        self.__dic_id_pos[str(idd)]["Windowid"],
                    )
            print("------------Fim-----------------------")

            for idd in self.__lst_VisiveisAnterior:
                if idd in self.__lst_NotVisiveis:
                    if self.__dic_id_pos[str(idd)]["Windowid"] != None:
                        print(idd, "destruidos")
                        self.__Canvas.delete(self.__dic_id_pos[str(idd)]["Windowid"])
                        self.__dic_id_pos[str(idd)]["Windowid"] = None

        def Atualizar_ScrolArea(self):
            self.__Canvas.update_idletasks()
            self.__Canvas.configure(
                scrollregion=(
                    0,
                    0,
                    self.__Canvas.winfo_width(),
                    self.__Element_Height * self.Get_NumeroElementos(),
                )
            )

        def Atualizar_Cores(self, idend):

            if self.__dic_id_pos[str(idend)]["Posid"] <= self.Get_NumeroElementos() - 1:

                for idd in self.__lst_OrdemDosEementos[
                    : self.__dic_id_pos[str(idend)]["Posid"]
                ]:

                    elemento = self.__dic_id_pos[str(idd)]["Elemento"]
                    if elemento.Get_Color() == self.__ColorClara:
                        elemento.Set_Color(self.__ColorEscura)
                    else:
                        elemento.Set_Color(self.__ColorClara)

        def MoverPara_Visivel(self, idd, mode, x, y, stepx=25, stepy=10):
            """
            Move um elemento da tebela por um incremento
            """
            stid = str(idd)
            if mode == "paraCima":

                to = self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1] - stepy
                if to > y:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0],
                        to,
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(y=to)
                    return False
                else:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0],
                        y,
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(y=y)
                    return True

            elif mode == "paraBaixo":

                to = self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1] + stepy
                if to < y:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0],
                        to,
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(y=to)
                    return False
                else:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0],
                        y,
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(y=y)
                    return True

            elif mode == "paraDireita":

                to = self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0] + stepx
                if to < x:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        to,
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1],
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(x=to)
                    return False
                else:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        x,
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1],
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(x=x)
                    return True

            elif mode == "paraEsquerda":

                to = self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1] - stepx
                if to > y:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        to,
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1],
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(y=to)
                    return False
                else:
                    self.__Canvas.coords(
                        self.__dic_id_pos[stid]["Windowid"],
                        x,
                        self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1],
                    )
                    self.__dic_id_pos[stid]["Elemento"].Set_Geometria(y=y)
                    return True
            else:
                return (None,)

        def MoverPara_NaoVisivel(self, idd, mode, x, y):
            """
            Move um elemento da tebela de imediato
            """
            stid = str(idd)
            if mode == "paraCima" or mode == "paraBaixo":
                self.__Canvas.coords(
                    self.__dic_id_pos[stid]["Windowid"],
                    self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0],
                    y,
                )
                self.__dic_id_pos[stid]["Elemento"].Set_Geometria(y=y)
                return True
            elif mode == "paraDireita" or mode == "paraEsquerda":
                self.__Canvas.coords(
                    self.__dic_id_pos[stid]["Windowid"],
                    x,
                    self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1],
                )
                self.__dic_id_pos[stid]["Elemento"].Set_Geometria(x=x)
                return True
            else:
                return None

        def Mover_Elemento(
            self, idd, mode, x, y, stepx=25, stepy=10, time=5, tag="SemTag", **kwargs
        ):

            kwargs = kwargs

            if idd in self.__lst_Visiveis:
                b1 = self.MoverPara_Visivel(idd, mode, x, y)

                if b1 == None:
                    pass
                elif not b1:
                    self.after(
                        time,
                        lambda idd=idd, mode=mode, x=x, y=y, stepx=stepx, stepy=stepy, time=time: self.Mover_Elemento(
                            idd, mode, x, y, stepx, stepy, time, tag=tag, **kwargs
                        ),
                    )
                elif b1 == True:
                    self.RunAnimeBindings(tag, **kwargs)

            else:
                b = self.MoverPara_NaoVisivel(idd, mode, x, y)
                if b == None:
                    pass
                elif b == True:
                    self.RunAnimeBindings(tag, **kwargs)

            return b1

        def Mover_Elementos(
            self,
            ids,
            mode,
            x,
            y,
            delta=False,
            stepx=25,
            stepy=10,
            time=10,
            tag="SemTag",
            _exe=0,
            _dic={},
            **kwargs
        ):

            kwargs = kwargs

            if delta:

                for idd in ids:
                    stid = str(idd)

                    if _exe == 0:

                        if idd in self.__lst_Visiveis:

                            _dic[stid] = (
                                self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0]
                                - x,
                                self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1]
                                - y,
                            )
                            b = self.MoverPara_Visivel(
                                idd,
                                mode,
                                _dic[stid][0],
                                _dic[stid][1],
                                stepx=stepx,
                                stepy=stepy,
                            )

                        else:
                            self.MoverPara_NaoVisivel(
                                idd,
                                mode,
                                self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[0]
                                - x,
                                self.__dic_id_pos[stid]["Elemento"].Get_Geometria()[1]
                                - y,
                            )

                    elif idd in self.__lst_Visiveis:
                        b = self.MoverPara_Visivel(
                            idd,
                            mode,
                            _dic[stid][0],
                            _dic[stid][1],
                            stepx=stepx,
                            stepy=stepy,
                        )

                if not b:
                    _exe += 1
                    self.after(
                        time,
                        lambda ids=ids: self.Mover_Elementos(
                            ids,
                            mode,
                            x,
                            y,
                            delta,
                            stepx,
                            stepy,
                            time,
                            _dic=_dic,
                            _exe=_exe,
                            tag=tag,
                            **kwargs
                        ),
                    )
                else:
                    self.RunAnimeBindings(tag, **kwargs)
            else:
                for idd in ids:

                    if idd in self.__lst_Visiveis:
                        b = self.MoverPara_Visivel(
                            idd, mode, x, y, stepx=stepx, stepy=stepy
                        )
                    elif _exe == 0:
                        self.MoverPara_NaoVisivel(
                            idd,
                            mode,
                            x,
                            y,
                        )

                if not b:
                    _exe += 1
                    self.after(
                        time,
                        lambda idd=ids: self.Mover_Elementos(
                            ids,
                            mode,
                            x,
                            y,
                            delta,
                            stepx,
                            stepy,
                            time,
                            _dic=_dic,
                            _exe=_exe,
                            tag=tag,
                            **kwargs
                        ),
                    )
                else:
                    self.RunAnimeBindings(tag, **kwargs)

        def Novo_Elemento(self, livro, idd=None):
            """
            Cria uma nova linha na tabela return o id do elemento livro.
            """
            if str(self.__Last_id) in self.__dic_id_pos:
                color = self.__dic_id_pos[str(self.__Last_id)]["Elemento"].Get_Color()

                if color == self.__ColorEscura:
                    color = self.__ColorClara
                elif color == self.__ColorClara:
                    color = self.__ColorEscura

            else:
                color = self.__ColorClara

            if idd == None:
                self.__Last_id += 10
            else:
                self.__Last_id = int(idd)

            self.__Canvas.update_idletasks()
            self.__focuspos += self.__Element_Height

            elemento = self.Elemento(
                Tabela=self,
                livro=livro,
                posx=0,
                posy=self.__focuspos - self.__Element_Height,
                color=color,
                ID=self.__Last_id,
            )

            elemento.ONResize()

            self.__lst_OrdemDosEementos += [self.__Last_id]
            idw = self.__Canvas.create_window(
                0, self.__focuspos - self.__Element_Height, window=elemento, anchor=NW
            )
            self.__dic_id_pos[str(self.__Last_id)] = {
                "Elemento": elemento,
                "Windowid": idw,
                "Posid": self.Get_NumeroElementos() - 1,
            }
            self.Atualizar_listas()

            # self.__Last_id += 10  TAVA AQUI

            self.__Ordenados = False

            self.Atualizar_ScrolArea()
            """
            self.__lst_OrdemDosEementos += [self.__Last_id]
            #idw = self.__Canvas.create_window(0,self.__focuspos-self.__Element_Height,window=elemento, anchor=NW)
            self.__Canvas.configure(scrollregion=(0,0,self.__Canvas.winfo_width(),self.__focuspos))
            self.__dic_id_pos[str(self.__Last_id)] = {"Elemento":elemento,"Windowid":None}
            self.Atualizar_visiveis()
            self.__Last_id += 10 
            return self.__Last_id
            """
            return self.__Last_id

        def Novos_Elementos(self, livro):
            """
            Cria uma nova linha para cada livro no argumento (lista/tuplo)livros:
            Retorna os respetivos ids
            """
            t = 20
            p = MyWidgets.Progress_Bar(self, 0, 5, steps=t, Text="Loading")
            for i in range(t):
                self.Novo_Elemento(livro)
                p.Progredir(text=str(i))

            self.__Ordenados = False

        def Remover_Elemento(self, idd, stx=25, sty=10, time=10):
            """
            Remove o elemento e toda a sua informação da tabela
            """
            if str(idd) in self.__dic_id_pos and not self.__animando:
                self.__animando = True
                self.CreateAnime(tag="Removido", func=self.__Acabar_de_remover)
                if idd in self.__lst_Visiveis:
                    self.Mover_Elemento(
                        idd=idd,
                        mode="paraDireita",
                        x=self.__dic_id_pos[str(idd)]["Elemento"].Get_Geometria()[2],
                        y=0,
                        tag="Removido",
                        **{"idds": idd, "stx": stx, "sty": sty, "time": time}
                    )
                else:
                    self.__Acabar_de_remover(idd)

                self.__Ordenados = False
            else:
                pass

        def Remover_Elementos(self, iterIds, stx=25, sty=10, time=10):
            """
            Remove os elementos todas a sua informação da tabela
            """
            if not self.__animando:
                abort = False

                for idd in iterIds:
                    if not str(idd) in self.__dic_id_pos:
                        abort = True

                if not abort:
                    self.__animando = True
                    self.CreateAnime(tag="Removido", func=self.__Acabar_de_remover)
                    self.Mover_Elementos(
                        ids=iterIds,
                        mode="paraDireita",
                        x=self.__dic_id_pos[str(idd)]["Elemento"].Get_Geometria()[2],
                        y=0,
                        tag="Removido",
                        **{"idds": iterIds, "stx": stx, "sty": sty, "time": time}
                    )
                    self.__Ordenados = False
                else:
                    pass

        def __Acabar_de_remover(self, idds, stx=25, sty=10, time=10):

            # Erro quando pede-se para apagar enquanto está a apagar

            if not isinstance(idds, tuple):
                t = ()
                did = str(idds)
                for i in range(
                    self.__dic_id_pos[did]["Posid"] + 1, self.Get_NumeroElementos()
                ):
                    t += (self.__lst_OrdemDosEementos[i],)

                self.__dic_id_pos[did]["Elemento"].delete_imagepath()
                del self.__lst_OrdemDosEementos[self.__dic_id_pos[did]["Posid"]]
                self.Atualizar_listas()
                self.__focuspos -= self.Get_Elemento_Height()

                if t != ():
                    self.Mover_Elementos(
                        ids=t,
                        mode="paraCima",
                        x=0,
                        y=self.__Element_Height,
                        delta=True,
                        stepx=stx,
                        stepy=sty,
                        time=time,
                    )

                self.__dic_id_pos[did]["Elemento"].destroy()
                self.__animando = False

                self.Atualizar_Cores(idds)
                self.Atualizar_ScrolArea()
                del self.__dic_id_pos[did]

                self.event_generate("<<ENDANIM>>", x=idds)

            else:
                t = ()
                did = str(idds[-1])
                for i in range(
                    self.__dic_id_pos[did]["Posid"] + 1, self.Get_NumeroElementos()
                ):
                    t += (self.__lst_OrdemDosEementos[i],)

                for Idd in idds:
                    self.__lst_OrdemDosEementos.remove(Idd)

                self.Atualizar_listas()

                if t != ():
                    self.Mover_Elementos(
                        ids=t,
                        mode="paraCima",
                        x=0,
                        y=self.__Element_Height * len(idds),
                        delta=True,
                        stepx=stx,
                        stepy=sty,
                        time=time,
                    )

                for Idd in idds:
                    self.__dic_id_pos[str(Idd)]["Elemento"].destroy()

                self.__animando = False

                self.Atualizar_Cores(idds[0])
                self.Atualizar_ScrolArea()

                for Idd in idds:
                    del self.__dic_id_pos[str(Idd)]

        def Ordenar(self, por):
            pass

        def Prucurar(self, por):
            pass

        def MoveViewTo(self, idd):

            if str(idd) in self.__dic_id_pos:
                self.__Canvas.update_idletasks()
                self.__Canvas.yview(
                    "moveto",
                    self.__dic_id_pos[str(idd)]["Elemento"].Get_Geometria()[1]
                    / (self.Get_NumeroElementos() * self.Get_Elemento_Height()),
                )

        def ScrollbarMove(self, x, y, number=None):

            if number == None:
                self.__Canvas.yview(x, y)
            else:
                self.__Canvas.yview(x, y, number)

            # self.Atualizar_visiveis()
            self.On_Resize()

        def On_MouseWheel(self, event):
            self.__Canvas.yview_scroll(-event.delta // 120, "units")
            self.On_Resize()

        def On_Resize(self, event=None):
            self.Atualizar_listas()
            for idd in self.__lst_Visiveis:
                self.__dic_id_pos[str(idd)]["Elemento"].ONResize()
                pass

        def CreateAnime(self, tag, func):

            if tag in self.__AnimationQueue:
                if self.__AnimationQueue[tag] == ():
                    self.__AnimationQueue[tag] = (func,)
                else:
                    self.__AnimationQueue[tag] += (func,)
            else:
                self.__AnimationQueue[tag] = (func,)

        def DeleteAnime(self, func):

            for key in self.__AnimationQueue:

                for funcao in self.__AnimationQueue[key]:

                    if funcao == func:
                        self.__AnimationQueue[key].remove(func)

        def RunAnimeBindings(self, tag, **kwargs):

            if tag in self.__AnimationQueue:
                for func in self.__AnimationQueue[tag]:

                    func(**kwargs)

                del self.__AnimationQueue[tag]

        class Elemento(Canvas):
            def __init__(self, Tabela, livro, posx, posy, color, ID):

                # livro.Set_Autor(str(ID))
                self.__tabela = Tabela
                # get master
                self.__Canvs = Tabela.Get_Canvas()
                self.__Canvs.update_idletasks()
                # info
                self.__Color = color
                self.__ColorSelected = "light blue"
                self.__Livro = livro
                self.__TxtsId_dic = {}
                self.__id = ID
                self.__Suspended = False
                self.__Selected = False
                self.__click = False
                # Dimenções
                self.__pad = 4
                self.__posx = posx
                self.__posy = posy
                self.__h = Tabela.Get_Elemento_Height()
                self.__w = 500  # self.__Canvs.winfo_width()-20

                super().__init__(
                    self.__Canvs,
                    width=self.__Canvs.winfo_width(),
                    height=80,
                    highlightthickness=0,
                    bg=self.__Color,
                )

                # elementos (Estado,image,infotable)
                self.__Rect_Estado_id = self.create_rectangle(
                    0,
                    0,
                    20,
                    self.__h,
                    tags=("Estado"),
                    fill="light green",
                    outline="",
                    width=0,
                )

                if livro.Get_ImagemPath() == None or livro.Get_ImagemPath() == "n/d":
                    self.__Rect_Image_id = self.create_rectangle(
                        0 + 20 + self.__pad,
                        0 + self.__pad,
                        120 - self.__pad,
                        0 + (self.__h - self.__pad),
                        tags=("Image",),
                        fill="Gray",
                        outline="",
                        width=0,
                    )
                    self.image_ori = None
                    self.image_in = None
                    self.img = None
                    self.imagemm = False
                else:
                    self.image_ori = Image.open(livro.Get_ImagemPath())
                    self.image_in = self.image_ori.resize(
                        (
                            100 - 2 * self.__pad,
                            self.__tabela.Get_Elemento_Height() - 2 * self.__pad,
                        )
                    )
                    self.img = ImageTk.PhotoImage(self.image_in)
                    self.__Rect_Image_id = self.create_image(
                        20 + self.__pad,
                        self.__pad,
                        anchor=NW,
                        image=self.img,
                        tags=("Image",),
                    )
                    self.imagemm = True

                self.Comfigurar(altura=self.__h)
                self.Refresh_Texts()

                self.bind("<Enter>", self.Mouse_in)
                self.bind("<Leave>", self.Mouse_out)
                self.bind("<Button>", self.Event_ONclick)

            def Get_id(self):
                return self.__id

            def Get_Color(self):
                return self.__Color

            def Get_livro(self):
                return self.__Livro

            def Get_State(self):
                return self.__Suspended

            def Set_State(self, state):
                self.__Suspended = state

            def Get_Geometria(self):
                """
                Retorna um tuplo (x,y,w,h)
                """
                return (self.__posx, self.__posy, self.__w, self.__h)

            def Change_State(self):
                self.__Suspended = not self.__Suspended

            def Set_Color(self, color):

                # verificar se quero resetar a cor seestiver selecionado

                self.__Color = color
                self.configure(
                    bg=self.__Color,
                    highlightthickness=0,
                    height=self.__h,
                    width=self.__w,
                )
                if self.__Selected:
                    self.__Selected = False
                    self.__click = False
                    self.__tabela.Set_Selected(None)

            def Set_Geometria(self, x=None, y=None, w=None, h=None):

                if x != None:
                    self.__posx = x

                if y != None:
                    self.__posy = y

                if w != None:
                    self.__w = w

                if h != None:
                    self.__h = h

            def Comfigurar(self, color=None, image=None, altura=None):

                if color != None:
                    self.configure(bg=color)

                if image != None and isinstance(image, str):

                    self.__Livro.Set_ImagemPath(image)
                    self.image_ori = Image.open(image)
                    self.image_in = self.image_ori.resize(
                        (
                            100 - 2 * self.__pad,
                            self.__tabela.Get_Elemento_Height() - 2 * self.__pad,
                        )
                    )
                    self.img = ImageTk.PhotoImage(self.image_in)
                    self.__Rect_Image_id = self.create_image(
                        20 + self.__pad,
                        self.__pad,
                        anchor=NW,
                        image=self.img,
                        tags=("Image",),
                    )

                if altura != None:
                    if altura >= 80:

                        self.delete("TEXTOS")
                        self.__h = int(altura)

                        tempdic = self.__Livro.Get_Basic_Info()
                        tempkeylist = list(tempdic.keys())
                        i = 0
                        lim = 0
                        stepy = Settings.Get("Tab_Font_2")["T"]

                        while (i == 0 or i < lim) and i < len(tempdic):

                            txtTITLE = self.create_text(
                                self.__posx + 130,
                                5 + i * stepy,
                                text=tempkeylist[i] + ":",
                                fill=Settings.Get("Tab_Font_2")["FG"],
                                font=(
                                    Settings.Get("Tab_Font_2")["Font"],
                                    Settings.Get("Tab_Font_2")["T"],
                                    "bold",
                                ),
                                justify=LEFT,
                                anchor=NW,
                                tags="TEXTOS",
                            )

                            dimentions = self.bbox(txtTITLE)
                            x = abs(dimentions[2] - dimentions[0]) + 5
                            y = abs(dimentions[3] - dimentions[1])

                            txtTEXT = self.create_text(
                                (self.__posx + 130 + x, 5 + i * stepy),
                                text=tempdic[tempkeylist[i]],
                                fill=Settings.Get("Tab_Font_2")["FG"],
                                font=(
                                    Settings.Get("Tab_Font_2")["Font"],
                                    Settings.Get("Tab_Font_2")["T"],
                                    "italic",
                                ),
                                justify=LEFT,
                                anchor=NW,
                                tags="TEXTOS",
                            )

                            self.__TxtsId_dic[tempkeylist[i]] = {
                                "Title": txtTITLE,
                                "Text": txtTEXT,
                            }

                            if i == 0:
                                stepy = y
                                lim = round((self.__h - 10) / stepy)

                            i += 1

                        self.configure(height=self.__h)
                        self.coords(
                            self.__Rect_Estado_id,
                            self.__posx,
                            0,
                            self.__posx + 20,
                            0 + self.__h,
                        )
                        if not self.imagemm:
                            self.coords(
                                self.__Rect_Image_id,
                                self.__posx + 20 + self.__pad,
                                0 + self.__pad,
                                self.__posx + 120 - self.__pad,
                                0 + (self.__h - self.__pad),
                            )

                        del (tempdic, tempkeylist, lim, stepy)

            def Refresh_Texts(self, font=None):

                self.__Canvs.update_idletasks()
                self.__w = self.__Canvs.winfo_width()
                dic = self.__Livro.Get_Basic_Info()

                if font != None:
                    self.Comfigurar(altura=self.__h)

                for key in self.__TxtsId_dic:

                    dimentions = self.bbox(self.__TxtsId_dic[key]["Text"])
                    x = abs(dimentions[2] - dimentions[0])
                    # print(x,self.itemcget(self.__TxtsId_dic[key]["Text"],"text"))

                    carateres = round(
                        (self.__w - 220)
                        / (
                            x
                            / len(self.itemcget(self.__TxtsId_dic[key]["Text"], "text"))
                        )
                    )

                    if carateres < len(dic[key]):
                        self.itemconfigure(
                            self.__TxtsId_dic[key]["Text"],
                            text=dic[key][: carateres - 3] + "...",
                        )
                    else:
                        self.itemconfigure(
                            self.__TxtsId_dic[key]["Text"], text=dic[key]
                        )

            def Mouse_in(self, event=None):
                if not self.__click:
                    self.itemconfigure(self.__Rect_Estado_id, fill="light green")
                    self.configure(
                        bg=self.__ColorSelected,
                        highlightthickness=3,
                        height=self.__h - 6,
                        width=self.__w - 6,
                    )
                    self.__Selected = True

            def Mouse_out(self, event=None):
                if not self.__click:
                    self.configure(
                        bg=self.__Color,
                        highlightthickness=0,
                        height=self.__h,
                        width=self.__w,
                    )
                    self.__Selected = False
                    self.__click = False

            def Event_ONclick(self, event=None):

                if not self.__tabela.Get_MultiSelect():

                    if self.__tabela.Get_Selected() not in (None, self.__id):
                        self.__tabela.Desselecionar(self.__tabela.Get_Selected())

                    if not self.__click:
                        self.__click = True
                        self.configure(bg="yellow")
                        self.configure(
                            bg="yellow",
                            highlightthickness=3,
                            height=self.__h - 6,
                            width=self.__w - 6,
                        )
                        self.__tabela.Set_Selected(self.__id)
                    else:
                        self.Click_Reset()
                else:
                    if not self.__click:
                        self.__click = True
                        self.configure(bg="yellow")
                        self.__tabela.Set_Selected(self.__id)
                    else:
                        self.Click_Reset()

            def Click_Reset(self):

                # if self.__tabela.Get_MultiSelect():
                #    self.__tabela.Desselecionar(idd=self.__id)

                self.__click = False
                self.configure(bg=self.__Color)
                self.configure(
                    bg=self.__Color,
                    highlightthickness=0,
                    height=self.__h,
                    width=self.__w,
                )
                self.__Selected = False

                self.__tabela.event_generate("<<Desselecionado>>", x=self.__id)

            def ONResize(self):

                self.__Canvs.update_idletasks()
                self.__w = self.__Canvs.winfo_width()
                if self.__Selected:
                    self.configure(width=self.__w - 6, height=self.__h - 6)
                else:
                    self.configure(width=self.__w, height=self.__h)
                self.Refresh_Texts()

            def delete_imagepath(self):

                if self.__Livro.Get_ImagemPath() not in (None, "n/d"):
                    PATHREMOVE(self.__Livro.Get_ImagemPath())

    class Frame_Adicionar(Frame):

        """
        Abre uma interface para inserir um novo livro
        """

        def __init__(self, master, tabela):

            super().__init__(master=master)
            self.configure(bg="White", height=300)

            self.__master = master
            self.__tabela = tabela

            # image vars
            self.path = None
            self.path2 = None
            self.image = None
            self.image_in = None
            self.img_image_in = None
            self.Id = None

            # colunas
            self.columnconfigure(1, minsize=80, weight=1)
            self.columnconfigure(
                (0, 3, 5, 6, 7, 10), minsize=10, weight=0
            )  # espaços brancos
            self.columnconfigure((2, 9), minsize=20, weight=0)
            self.columnconfigure(4, minsize=80, weight=1)
            self.columnconfigure(8, minsize=40, weight=1)

            # linhas
            self.rowconfigure((0, 25), minsize=5, weight=0)
            self.rowconfigure(
                (2, 5, 8, 11, 14, 17, 20, 25, 27), minsize=5, weight=0
            )  # espaços brancos
            self.rowconfigure(23, minsize=15, weight=0)  # espaços brancos
            self.rowconfigure(
                (4, 7, 10, 13, 16, 19, 22, 26), minsize=25, weight=0
            )  # Entry's
            self.rowconfigure(24, minsize=25, weight=1)  # espaços brancos

            fonttitles = (
                Settings.Get("Tab_Font_1")["Font"],
                Settings.Get("Tab_Font_1")["T"],
                "italic",
            )
            fonttext = (
                Settings.Get("Tab_Font_2")["Font"],
                Settings.Get("Tab_Font_1")["T"],
                "bold",
            )
            textbgcolor = Settings.Get("Tab_Font_1")["FG"]

            self.__lb_FrameTitle = Label(
                self,
                text="Adicionar Livro",
                font=(
                    Settings.Get("Tab_Font_1")["Font"],
                    round(Settings.Get("Tab_Font_1")["T"] * 1.5),
                ),
                bg="White",
                fg=Settings.Get("Tab_Font_1")["FG"],
            )
            self.__lb_FrameTitle.grid(column=1, row=1, columnspan=5, sticky=NW)

            self.__lb_nome = Label(
                self, text="Nome:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_nome.grid(column=1, row=3, columnspan=9, sticky=SW)
            self.__tbox_nome = Entry(
                self,
                justify="center",
                bg="#d9d9d9",
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_nome.bind(
                "<KeyPress>",
                lambda event: self.__tbox_nome.configure(
                    highlightcolor=None, highlightthickness=0
                ),
            )
            self.__tbox_nome.grid(column=1, row=4, columnspan=9, sticky=NSEW)

            self.__lb_autor = Label(
                self, text="Autor:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_autor.grid(column=1, row=6, columnspan=9, sticky=SW)
            self.__tbox_autor = Entry(
                self,
                justify="center",
                bg="#d9d9d9",
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_autor.bind(
                "<KeyPress>",
                lambda event: self.__tbox_autor.configure(
                    highlightcolor=None, highlightthickness=0
                ),
            )
            self.__tbox_autor.grid(column=1, row=7, columnspan=9, sticky=NSEW)

            self.__lb_editora = Label(
                self, text="Editora:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_editora.grid(column=1, columnspan=9, row=9, sticky=SW)
            self.__tbox_editora = Entry(
                self,
                justify="center",
                bg="#d9d9d9",
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_editora.bind(
                "<KeyPress>",
                lambda event: self.__tbox_editora.configure(
                    highlightcolor=None, highlightthickness=0
                ),
            )
            self.__tbox_editora.grid(column=1, columnspan=9, row=10, sticky=NSEW)

            self.__lb_genero = Label(
                self, text="Género:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_genero.grid(column=1, columnspan=2, row=12, sticky=SW)
            self.__tbox_genero = Entry(
                self,
                justify="center",
                bg="#d9d9d9",
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_genero.bind(
                "<KeyPress>",
                lambda event: self.__tbox_genero.configure(
                    highlightcolor=None, highlightthickness=0
                ),
            )
            self.__tbox_genero.grid(column=1, columnspan=2, row=13, sticky=NSEW)

            self.__lb_numero = Label(
                self, text="Número:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_numero.grid(column=4, columnspan=3, row=12, sticky=SW)
            self.__tbox_numero = Entry(
                self,
                justify="center",
                bg="#d9d9d9",
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_numero.bind(
                "<KeyPress>",
                lambda event: self.Validar_numero(
                    event=event, entry=self.__tbox_numero
                ),
            )
            self.__tbox_numero.bind(
                "<KeyRelease>",
                lambda event: self.Validar_numero(
                    event=event, entry=self.__tbox_numero
                ),
            )
            self.__tbox_numero.grid(column=4, columnspan=3, row=13, sticky=NSEW)

            self.__lb_edicao = Label(
                self, text="Edição:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_edicao.grid(column=8, row=12, columnspan=2, sticky=SW)
            self.__combo_edicao = ttk.Combobox(
                self,
                values=list(range(1, 100)),
                background="#d9d9d9",
                state="readonly",
                font=fonttext,
            )
            self.__combo_edicao.current(0)
            self.__combo_edicao.grid(column=8, columnspan=2, row=13)

            self.__lb_condicao = Label(
                self, text="Condição:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_condicao.grid(column=1, columnspan=2, row=15, sticky=SW)
            self.__combo_condicao = ttk.Combobox(
                self,
                values=["Mt.Baixa", "Baixa", "Média", "Boa", "Mt. Boa", "Exelente"],
                background="#d9d9d9",
                state="readonly",
                font=fonttext,
            )
            self.__combo_condicao.current(3)
            self.__combo_condicao.grid(column=1, columnspan=2, row=16)

            self.__lb_datalan = Label(
                self, text="Lançamento:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_datalan.grid(column=1, row=18, sticky=SW)
            self.__tbox_datalan = Entry(
                self,
                justify="center",
                bg="#d9d9d9",
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_datalan.insert(0, "mm/dd/aa")
            self.__tbox_datalan.grid(column=1, row=19, sticky=NSEW)
            self.__tbox_datalan.bind(
                "<KeyPress>",
                lambda event: self.__tbox_datalan.configure(
                    highlightcolor=None, highlightthickness=0
                ),
            )
            Button(
                self,
                text="\/",
                activebackground="light blue",
                command=self.On_Click_datalan,
            ).grid(column=2, row=19, sticky=NSEW)
            self.__bool_lan = False

            self.__lb_dataReci = Label(
                self, text="Receção:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_dataReci.grid(column=1, row=21, sticky=SW)
            self.__tbox_dataReci = Entry(
                self,
                justify="center",
                bg="#d9d9d9",
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_dataReci.bind(
                "<KeyPress>",
                lambda event: self.__tbox_dataReci.configure(
                    highlightcolor=None, highlightthickness=0
                ),
            )
            self.__tbox_dataReci.insert(0, "mm/dd/aa")
            self.__tbox_dataReci.grid(column=1, row=22, sticky=NSEW)
            Button(
                self,
                text="\/",
                activebackground="light blue",
                command=self.On_Click_dataReci,
            ).grid(column=2, row=22, sticky=NSEW)
            self.__bool_rec = False

            self.__lb_outros = Label(
                self,
                text="Outras Informações:",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
            )
            self.__lb_outros.grid(column=1, row=23, columnspan=8, sticky=SW)
            self.__tbox_outros = Text(self, bg="#d9d9d9", relief=SUNKEN, font=fonttext)
            self.__tbox_outros.grid(column=1, row=24, columnspan=8, sticky=NSEW)
            self.__Scrollbar = Scrollbar(self, command=self.__tbox_outros.yview)
            self.__Scrollbar.grid(column=9, row=24, sticky=NSEW)
            self.__tbox_outros.configure(yscrollcommand=self.__Scrollbar.set)

            self.__bt_adicionar = Button(
                self, text="Adicionar", command=self.On_click_Adicionar
            )
            self.__bt_adicionar.grid(column=6, columnspan=4, row=26, sticky=NSEW)

            self.__bt_cancelar = Button(
                self, text="Cancelar", command=self.On_click_Cancelar
            )
            self.__bt_cancelar.grid(column=4, columnspan=1, row=26, sticky=NSEW)

            f = Frame(
                self, relief=RAISED, highlightthickness=5, highlightbackground="white"
            )
            f.grid(column=4, columnspan=6, row=15, rowspan=7, sticky=NSEW)
            self.__Canvas = Canvas(f, width=100, height=100)
            self.__Canvas.pack(anchor=CENTER)

            self.__bt_imagem = Button(
                self, text="Selecionar", command=self.On_click_Selecionar
            )
            self.__bt_imagem.grid(column=4, columnspan=6, row=22, sticky=NSEW)

            # bindings
            self.bind("<Button-1>", lambda event: self.focus_set())
            self.bind("<Button-2>", lambda event: self.focus_set())
            self.bind("<Button-3>", lambda event: self.focus_set())

            self.myedit = None

            # combobox
            if isinstance(tabela, list) or isinstance(tabela, tuple):

                self.__combobox = ttk.Combobox(
                    self, values=["Tabela"], background="#d9d9d9", state="readonly"
                )
                self.rowconfigure(2, minsize=20, weight=0)
                self.__combobox.current(0)
                self.__combobox.grid(column=1, columnspan=2, row=2, sticky=NSEW)

            # datepiker
            self.__frm_calendario = Frame(self, bg="white")
            self.__frm_calendario.rowconfigure(0, weight=1)
            self.__frm_calendario.rowconfigure(1, minsize=20, weight=0)
            self.__frm_calendario.columnconfigure(0, weight=1)
            self.__time = datetime.now()
            self.__calendario = Calendar(
                self.__frm_calendario,
                selectmode="day",
                year=int(self.__time.strftime("%y")),
                month=int(self.__time.strftime("%m")),
                day=int(self.__time.strftime("%d")),
            )
            self.__calendario.grid(row=0, column=0, columnspan=3, sticky=NSEW)
            Button(
                self.__frm_calendario, text="Concluido", command=self.On_Click_Concluido
            ).grid(row=1, column=0, columnspan=2, sticky=NSEW)

        """
        def Bind(self,event,func):
            
            if event in self.__events:
                self.__events[event] += [func]
                
        def UnBind(self,func):
            
            for event in self.__events:
                
                for func2 in self.__events[event]:
                    
                    if func2 == func:
                        self.__events[event].remove(func)
                        break
        """

        def Atualizar_entrys(sefl, entry):

            entry.configure(highlightcolor=None, highlightthickness=0)

        def Validar_numero(self, event, entry):

            self.Atualizar_entrys(entry)

            s = entry.get()
            if len(s) == 1:
                if s[0] not in ("0", "1", "2", "3", "4", "5", "6", "7", "8", "9"):
                    entry.delete(0, END)
            elif len(s) > 0 and s[-1] not in (
                "0",
                "1",
                "2",
                "3",
                "4",
                "5",
                "6",
                "7",
                "8",
                "9",
            ):
                entry.delete(0, END)
                entry.insert(0, s[:-1])

        def On_Click_Concluido(self):

            if self.__bool_lan:
                self.__tbox_datalan.delete(0, END)
                self.__tbox_datalan.insert(0, self.__calendario.get_date())
                self.__tbox_datalan.configure(highlightcolor=None, highlightthickness=0)
                self.__frm_calendario.place_forget()
                self.__bool_lan = False
            else:
                self.__tbox_dataReci.delete(0, END)
                self.__tbox_dataReci.insert(0, self.__calendario.get_date())
                self.__tbox_dataReci.configure(
                    highlightcolor=None, highlightthickness=0
                )
                self.__frm_calendario.place_forget()
                self.__bool_rec = False

        def On_Click_datalan(self):

            if not self.__bool_lan:
                self.__calendario.selection_set(
                    "{}/{}/{}".format(
                        int(self.__time.strftime("%m")),
                        int(self.__time.strftime("%d")),
                        int(self.__time.strftime("%y")),
                    )
                )
                x = self.__tbox_datalan.winfo_x()
                y = self.__tbox_datalan.winfo_y()
                h = self.__tbox_datalan.winfo_height()

                self.__bool_rec = False

                self.update_idletasks()
                if self.winfo_height() - y < 300:
                    self.__frm_calendario.place(x=x, y=y - 235)
                else:
                    self.__frm_calendario.place(x=x, y=y + h)
            else:
                self.__frm_calendario.place_forget()

            self.__bool_lan = not self.__bool_lan

        def On_Click_dataReci(self):

            if not self.__bool_rec:
                self.__calendario.selection_set(
                    "{}/{}/{}".format(
                        int(self.__time.strftime("%m")),
                        int(self.__time.strftime("%d")),
                        int(self.__time.strftime("%y")),
                    )
                )
                x = self.__tbox_dataReci.winfo_x()
                y = self.__tbox_dataReci.winfo_y()
                h = self.__tbox_dataReci.winfo_height()

                self.__bool_lan = False

                self.update_idletasks()
                if self.winfo_height() - y < 300:
                    self.__frm_calendario.place(x=x, y=y - 235)
                else:
                    self.__frm_calendario.place(x=x, y=y + h)
            else:
                self.__frm_calendario.place_forget()

            self.__bool_rec = not self.__bool_rec

        def On_click_Selecionar(self):

            if MyWidgets.EditImage.Counter < 1:

                self.path = filedialog.askopenfile(
                    mode="r",
                    title="Selecionar imagem",
                    filetypes=(
                        ("PNG files", "*.png"),
                        ("JPEG files", (".jpg", ".jpeg")),
                    ),
                )

                if self.path != None:
                    self.myedit = MyWidgets.EditImage(
                        self, self.path.name, self.image_add
                    )

                if isinstance(self.__master, Toplevel):
                    self.__master.attributes("-topmost", True)
                    self.__master.attributes("-topmost", False)

            else:
                messagebox.showinfo(
                    "Janela já aberta",
                    "Já está uma janela aberta,\nfeche ou cancele para poder abrir uma nova",
                )

        def image_add(self, image_opentype):

            self.image = image_opentype
            self.image_in = self.image.resize((150, 150))
            self.img_image_in = ImageTk.PhotoImage(self.image_in)
            self.Id = self.__Canvas.create_image(
                50, 50, anchor=CENTER, image=self.img_image_in
            )

            if self.myedit != None:
                self.myedit.cancelar()

        def On_click_Adicionar(self):

            self.__bt_adicionar.configure(state="disable")
            self.__bt_cancelar.configure(state="disable")

            string = ""
            c = 0
            dataL = self.__tbox_datalan.get().strip().split("/")
            dataR = self.__tbox_dataReci.get().strip().split("/")

            if isinstance(self.__tabela, list) or isinstance(self.__tabela, tuple):
                if self.__combobox.current() == 0:
                    self.__combobox.configure(foreground="red")
                    messagebox.showwarning(
                        "Aviso",
                        "Não é posivel adicionar um livro sem uma tabela selecionada.",
                    )
                    self.__bt_adicionar.configure(state="normal")
                    self.__bt_cancelar.configure(state="normal")

                    if isinstance(self.__master, Toplevel):
                        self.__master.attributes("-topmost", True)
                        self.__master.attributes("-topmost", False)

            elif self.__tbox_nome.get().strip(" ") == "":
                self.__tbox_nome.configure(
                    highlightthickness=2, highlightbackground="red"
                )
                messagebox.showwarning(
                    "Aviso", "Não é posivel adicionar um livro sem um nome."
                )
                self.__bt_adicionar.configure(state="normal")
                self.__bt_cancelar.configure(state="normal")

                if isinstance(self.__master, Toplevel):
                    self.__master.attributes("-topmost", True)
                    self.__master.attributes("-topmost", False)

            elif (
                not self.__tbox_numero.get().isnumeric()
                or self.__tbox_numero.get().strip(" ") == ""
            ):
                self.__tbox_numero.configure(
                    highlightthickness=2, highlightbackground="red"
                )
                messagebox.showwarning(
                    "Aviso", "Não é posivel adicionar um livro sem um Número"
                )
                self.__bt_adicionar.configure(state="normal")
                self.__bt_cancelar.configure(state="normal")

                if isinstance(self.__master, Toplevel):
                    self.__master.attributes("-topmost", True)
                    self.__master.attributes("-topmost", False)

            else:
                if len(dataL) == 3 and len(dataR) == 3:

                    cc = 0
                    for i in (2, 0, 1):

                        try:
                            dataL[i] = int(dataL[i])
                        except:
                            messagebox.showwarning(
                                "Aviso", "Data de lançamento está inválido"
                            )
                            self.__tbox_datalan.configure(
                                highlightthickness=2, highlightbackground="red"
                            )
                            break

                        try:
                            dataR[i] = int(dataR[i])
                        except:
                            messagebox.showwarning(
                                "Aviso", "Data de recibimento está inválido"
                            )
                            self.__tbox_dataReci.configure(
                                highlightthickness=2, highlightbackground="red"
                            )
                            break

                        if dataL[i] > dataR[i]:
                            messagebox.showwarning(
                                "Aviso",
                                "A data de lançamento ocorre depois da data de receção",
                            )
                            self.__tbox_datalan.configure(
                                highlightthickness=2, highlightbackground="red"
                            )
                            self.__tbox_dataReci.configure(
                                highlightthickness=2, highlightbackground="red"
                            )
                            break
                        else:
                            cc += 1

                    if cc == 3:
                        bb = True
                    else:
                        bb = False

                    self.__bt_adicionar.configure(state="normal")
                    self.__bt_cancelar.configure(state="normal")

                elif len(dataL) != 3:
                    messagebox.showwarning("Aviso", "Data de lançamento está inválido")
                    self.__tbox_datalan.configure(
                        highlightthickness=2, highlightbackground="red"
                    )
                    self.__bt_adicionar.configure(state="normal")
                    self.__bt_cancelar.configure(state="normal")

                elif len(dataR) != 3:
                    messagebox.showwarning("Aviso", "Data de recibimento está inválido")
                    self.__tbox_dataReci.configure(
                        highlightthickness=2, highlightbackground="red"
                    )
                    self.__bt_adicionar.configure(state="normal")
                    self.__bt_cancelar.configure(state="normal")

                if bb:

                    if self.__tbox_autor.get().strip(" ") == "":
                        self.__tbox_autor.configure(
                            highlightthickness=2, highlightbackground="red"
                        )
                        if not string == "":
                            string += "\n"
                        string += '"Autor" está em branco;'
                        c += 1
                    if self.__tbox_editora.get().strip(" ") == "":
                        self.__tbox_editora.configure(
                            highlightthickness=2, highlightbackground="red"
                        )
                        if not string == "":
                            string += "\n"
                        string += '"Editora" está em branco;'
                        c += 1
                    if self.__tbox_genero.get().strip(" ") == "":
                        self.__tbox_genero.configure(
                            highlightthickness=2, highlightbackground="red"
                        )
                        if not string == "":
                            string += "\n"
                        string += '"Género" está em branco;'
                        c += 1

                    if c > 0:
                        if messagebox.askyesno(
                            "Aviso",
                            "O(s) campo(s):\n"
                            + string
                            + "\nDeseja adicionar mesmo assim ?",
                            parent=self.__master,
                        ):

                            if self.__tbox_autor.get() == "":
                                self.__tbox_autor.insert(0, "n/d")

                            if self.__tbox_genero.get() == "":
                                self.__tbox_genero.insert(0, "n/d")

                            if self.__tbox_editora.get() == "":
                                self.__tbox_editora.insert(0, "n/d")

                            idd = self.__tabela.Novo_Elemento(
                                Livro(
                                    self.__tbox_numero.get(),
                                    self.__tbox_nome.get(),
                                    self.__tbox_autor.get(),
                                    self.__tbox_editora.get(),
                                    self.__tbox_genero.get(),
                                    self.__combo_edicao.get(),
                                    self.__tbox_dataReci.get(),
                                    self.__tbox_datalan.get(),
                                    self.__combo_condicao.get(),
                                    None,
                                    self.__tbox_outros.get("1.0", END),
                                )
                            )
                            if self.image != None:
                                self.path2 = "Livros.imagens/img" + str(idd) + ".png"
                                self.image.save(self.path2, "png")
                                self.__tabela.Get_Elementos(idd, "Elemento").Comfigurar(
                                    image=self.path2
                                )

                            self.__master.event_generate("<<click_adicionar>>")

                            self.destroy()

                            # numero,title,autor="n/d",editora="n/d",genero="n/d",edition="n/d",datain="n/d",dataout="n/d",condition="Bom Estado",imagem="n/d",outros = ,hist=Stack(),histp=Stack()):

                        else:

                            if isinstance(self.__master, Toplevel):
                                self.__master.attributes("-topmost", True)
                                self.__master.attributes("-topmost", False)

                            self.__bt_adicionar.configure(state="normal")
                            self.__bt_cancelar.configure(state="normal")
                    else:

                        idd = self.__tabela.Novo_Elemento(
                            Livro(
                                self.__tbox_numero.get(),
                                self.__tbox_nome.get(),
                                self.__tbox_autor.get(),
                                self.__tbox_editora.get(),
                                self.__tbox_genero.get(),
                                self.__combo_edicao.get(),
                                self.__tbox_dataReci.get(),
                                self.__tbox_datalan.get(),
                                self.__combo_condicao.get(),
                                None,
                                self.__tbox_outros.get("1.0", END),
                            )
                        )
                        if self.image != None:
                            self.path2 = "Livros.imagens/img" + str(idd) + ".png"
                            self.image.save(self.path2, "png")
                            self.__tabela.Get_Elementos(idd, "Elemento").Comfigurar(
                                image=self.path2
                            )

                        self.__master.event_generate("<<click_adicionar>>")

                        self.destroy()

        def On_click_Cancelar(self):

            """
            for func in self.__events["<Cancel>"]:
                func()
            """

            if isinstance(self.__master, Tk) or isinstance(self.__master, Toplevel):
                self.__master.destroy()
            else:
                self.destroy()

    class Frame_FichaTecnica(Frame):

        """
        Cria uma nova interfa para mostrar a ficha tecnica do livro
        """

        def __init__(self, master, tabela, livro=None):

            super().__init__(master=master, bg="white")

            self.__master = master
            self.__tabela = tabela
            self.__livro = livro

            # colunas
            self.columnconfigure((0, 2, 4, 7), minsize=10, weight=0)  # espaços brancos
            self.columnconfigure(1, minsize=100, weight=1)
            self.columnconfigure(3, minsize=80, weight=1)
            self.columnconfigure(5, minsize=60, weight=1)
            self.columnconfigure(6, minsize=20, weight=0)

            # linhas
            self.rowconfigure(
                (0, 2, 5, 8, 11, 14, 17, 20, 23, 26, 28), minsize=5, weight=0
            )  # espaços brancos
            self.rowconfigure((4, 7, 10, 13, 16, 24), minsize=25, weight=0)  # Entry's
            self.rowconfigure(19, minsize=50, weight=1, uniform="Row")
            self.rowconfigure(22, minsize=60, weight=1, uniform="Row")
            self.rowconfigure(25, minsize=60, weight=1, uniform="Row")

            fonttitles = (
                Settings.Get("Tab_Font_1")["Font"],
                Settings.Get("Tab_Font_1")["T"],
                "italic",
            )
            fonttext = (
                Settings.Get("Tab_Font_2")["Font"],
                Settings.Get("Tab_Font_1")["T"],
            )
            textbgcolor = Settings.Get("Tab_Font_1")["FG"]
            textboxbgcolor = "#d9d9d9"

            self.__lb_FrameTitle = Label(
                self,
                text="Ficha Técnica",
                font=(
                    Settings.Get("Tab_Font_1")["Font"],
                    round(Settings.Get("Tab_Font_1")["T"] * 1.5),
                ),
                bg="White",
                fg=Settings.Get("Tab_Font_1")["FG"],
            )
            self.__lb_FrameTitle.grid(column=1, row=1, columnspan=5, sticky=NW)

            # if isinstance(tabela,list) or isinstance(tabela,tuple):

            self.__bt_editar = Button(self, text="Editar", command=self.On_click_self)
            self.__bt_editar.grid(column=5, columnspan=2, row=27, sticky=NSEW)

            #    self.__combo_tabelas = ttk.Combobox(self,values=["Mt.Baixa","Baixa","Média","Boa","Mt. Boa","Exelente"],background="#d9d9d9",state="readonly",font=fonttext)
            #    self.__combo_tabelas.grid(column=1,row=2,columnspan=1,sticky=NSEW)

            #    self.__combo_livro = ttk.Combobox(self,values=["Mt.Baixa","Baixa","Média","Boa","Mt. Boa","Exelente"],background="#d9d9d9",state="readonly",font=fonttext)
            #    self.__combo_livro.grid(column=3,row=2,columnspan=1,sticky=NSEW)

            self.__lb_nome = Label(
                self, text="Nome:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_nome.grid(column=1, row=3, columnspan=6, sticky=SW)
            self.__tbox_nome = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_nome.insert(0, livro.Get_Title())
            self.__tbox_nome.configure(state="readonly")
            self.__tbox_nome.grid(column=1, row=4, columnspan=6, sticky=NSEW)

            self.__lb_autor = Label(
                self, text="Autor:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_autor.grid(column=1, row=6, columnspan=6, sticky=SW)
            self.__tbox_autor = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_autor.insert(0, livro.Get_Autor())
            self.__tbox_autor.configure(state="readonly")
            self.__tbox_autor.grid(column=1, row=7, columnspan=6, sticky=NSEW)

            self.__lb_editora = Label(
                self, text="Editora:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_editora.grid(column=1, columnspan=6, row=9, sticky=SW)
            self.__tbox_editora = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_editora.insert(0, livro.Get_Editora())
            self.__tbox_editora.configure(state="readonly")
            self.__tbox_editora.grid(column=1, columnspan=6, row=10, sticky=NSEW)

            self.__lb_genero = Label(
                self, text="Género:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_genero.grid(column=1, columnspan=1, row=12, sticky=SW)
            self.__tbox_genero = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_genero.insert(0, livro.Get_Genero())
            self.__tbox_genero.configure(state="readonly")
            self.__tbox_genero.grid(column=1, columnspan=1, row=13, sticky=NSEW)

            self.__lb_numero = Label(
                self, text="Número:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_numero.grid(column=3, columnspan=1, row=12, sticky=SW)
            self.__tbox_numero = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_numero.insert(0, livro.Get_Numero())
            self.__tbox_numero.configure(state="readonly")
            self.__tbox_numero.grid(column=3, columnspan=1, row=13, sticky=NSEW)

            self.__lb_edicao = Label(
                self, text="Edição:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_edicao.grid(column=5, row=12, columnspan=2, sticky=SW)
            self.__tbox_edicao = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_edicao.insert(0, livro.Get_Edition() + "º Edição")
            self.__tbox_edicao.configure(state="readonly")
            self.__tbox_edicao.grid(column=5, columnspan=2, row=13, sticky=NSEW)

            self.__lb_condicao = Label(
                self, text="Condição:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_condicao.grid(column=1, columnspan=1, row=15, sticky=SW)
            self.__tbox_condicao = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_condicao.insert(0, livro.Get_Condition())
            self.__tbox_condicao.configure(state="readonly")
            self.__tbox_condicao.grid(column=1, columnspan=1, row=16, sticky=NSEW)

            self.__lb_datalan = Label(
                self, text="Lançamento:", font=fonttitles, bg="White", fg=textbgcolor
            )
            self.__lb_datalan.grid(column=3, row=15, sticky=SW)
            self.__tbox_datalan = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_datalan.insert(0, livro.Get_DataOut())
            self.__tbox_datalan.configure(state="readonly")
            self.__tbox_datalan.grid(column=3, columnspan=1, row=16, sticky=NSEW)

            self.__lb_dataReci = Label(
                self,
                text="Receção:",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
                justify=CENTER,
            )
            self.__lb_dataReci.grid(column=5, columnspan=2, row=15, sticky=SW)
            self.__tbox_dataReci = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_dataReci.insert(0, livro.Get_DataIn())
            self.__tbox_dataReci.configure(state="readonly")
            self.__tbox_dataReci.grid(column=5, columnspan=2, row=16, sticky=NSEW)

            self.__lb_outros = Label(
                self,
                text="Outras Informações:",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
            )
            self.__lb_outros.grid(column=1, row=18, columnspan=6, sticky=SW)
            self.__listBox_info = Listbox(self, bg=textboxbgcolor)
            self.__listBox_info.grid(column=1, columnspan=5, row=19, sticky=NSEW)
            self.__Scrollbar3 = Scrollbar(self, command=self.__listBox_info.yview)
            self.__Scrollbar3.grid(column=6, row=19, sticky=NSEW)
            self.__listBox_info.configure(yscrollcommand=self.__Scrollbar3.set)

            self.__lb_Alugado = Label(
                self,
                text="Ultimos empréstimos:",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
            )
            self.__lb_Alugado.grid(column=1, row=21, columnspan=6, sticky=SW)
            self.__tv_emprestimos = ttk.Treeview(
                self, selectmode=BROWSE, columns=(1, 2, 3), show="headings"
            )
            self.__tv_emprestimos.heading(1, text="Desde")
            self.__tv_emprestimos.column(1, stretch=False, width=85, anchor=CENTER)
            self.__tv_emprestimos.heading(2, text="Até")
            self.__tv_emprestimos.column(2, stretch=False, width=85, anchor=CENTER)
            self.__tv_emprestimos.heading(3, text="Nome")
            self.__tv_emprestimos.column(3, stretch=True, width=85)
            self.__tv_emprestimos.grid(row=22, column=1, columnspan=5, sticky=NSEW)

            self.__Scrollbar2 = Scrollbar(self, command=self.__tv_emprestimos.yview)
            self.__Scrollbar2.grid(column=6, row=22, sticky=NSEW)
            self.__tv_emprestimos.configure(yscrollcommand=self.__Scrollbar2.set)
            self.__tv_emprestimos.insert(
                parent="", index=0, iid=0, values=("vineet", "e11", 1000000.00)
            )
            self.__tv_emprestimos.insert(
                parent="", index=1, iid=1, values=("anil", "e12", 120000.00)
            )
            self.__tv_emprestimos.insert(
                parent="", index=2, iid=2, values=("ankit", "e13", 41000.00)
            )
            self.__tv_emprestimos.insert(
                parent="", index=3, iid=3, values=("Shanti", "e14", 22000.00)
            )

            self.__lb_Alugado = Label(
                self,
                text="Registro de danos",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
            )
            self.__lb_Alugado.grid(column=1, row=24, columnspan=6, sticky=SW)
            self.__listBox = Listbox(self, bg=textboxbgcolor)
            self.__listBox.grid(column=1, columnspan=5, row=25, sticky=NSEW)
            self.__Scrollbar4 = Scrollbar(self, command=self.__listBox.yview)
            self.__Scrollbar4.grid(column=6, row=25, sticky=NSEW)
            self.__listBox.configure(yscrollcommand=self.__Scrollbar4.set)

            # bindings
            self.bind("<Button-1>", self.On_click_self)
            self.bind("<Button-2>", self.On_click_self)
            self.bind("<Button-3>", self.On_click_self)

        def On_click_self(self, event=None):

            if len(self.__tv_emprestimos.selection()) > 0:
                self.__tv_emprestimos.selection_remove(
                    self.__tv_emprestimos.selection()[0]
                )
            self.focus_set()

    class Frame_InformaçõesTabela(Frame):

        """
        Cria uma nova interface para mostrar as informações sobre a tabela aberta;
        """

        def __init__(self, master, tabela):

            super().__init__(master, bg="white", borderwidth=0)

            self.tabela = tabela

            # colunas
            self.columnconfigure((0, 3), minsize=10, weight=0)  # espaços brancos
            self.columnconfigure(1, minsize=260, weight=1)
            self.columnconfigure(2, minsize=20, weight=0)

            # linhas
            self.rowconfigure((0, 14), minsize=5, weight=0)  # espaços brancos
            self.rowconfigure((3, 6, 9, 12), minsize=25, weight=0)  # Entry's
            self.rowconfigure(13, minsize=50, weight=1, uniform="Row")
            self.rowconfigure(10, minsize=50, weight=1, uniform="Row")

            fonttitles = (
                Settings.Get("Tab_Font_1")["Font"],
                Settings.Get("Tab_Font_1")["T"],
                "italic",
            )
            fonttext = (
                Settings.Get("Tab_Font_2")["Font"],
                Settings.Get("Tab_Font_1")["T"],
            )
            textbgcolor = Settings.Get("Tab_Font_1")["FG"]
            textboxbgcolor = "#d9d9d9"

            self.__lb_FrameTitle = Label(
                self,
                text="Informações",
                font=(
                    Settings.Get("Tab_Font_1")["Font"],
                    round(Settings.Get("Tab_Font_1")["T"] * 1.5),
                ),
                bg="White",
                fg=Settings.Get("Tab_Font_1")["FG"],
            )
            self.__lb_FrameTitle.grid(column=1, row=1, columnspan=2, sticky=NW)

            self.__lb_nome = Label(
                self,
                text="Nome:",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
                justify=CENTER,
            )
            self.__lb_nome.grid(column=1, columnspan=2, row=3, sticky=SW)
            self.__tbox_nome = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_nome.insert(0, self.tabela.Get_Nome())
            self.__tbox_nome.configure(state="readonly")
            self.__tbox_nome.grid(column=1, columnspan=2, row=4, sticky=NSEW)

            self.__lb_nLivros = Label(
                self,
                text="Numero de livros:",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
                justify=CENTER,
            )
            self.__lb_nLivros.grid(column=1, columnspan=2, row=6, sticky=SW)
            self.__tbox_nLivros = Entry(
                self,
                justify="center",
                bg=textboxbgcolor,
                readonlybackground=textboxbgcolor,
                relief=SUNKEN,
                exportselection=0,
                font=fonttext,
            )
            self.__tbox_nLivros.insert(0, self.tabela.Get_NumeroElementos())
            self.__tbox_nLivros.configure(state="readonly")
            self.__tbox_nLivros.grid(column=1, columnspan=2, row=7, sticky=NSEW)

            self.__lb_outros = Label(
                self,
                text="Ultimos adicionados",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
            )
            self.__lb_outros.grid(column=1, row=9, columnspan=2, sticky=SW)
            self.__listBox_info = Listbox(self, bg=textboxbgcolor)
            self.__listBox_info.grid(column=1, row=10, sticky=NSEW)
            self.__Scrollbar3 = Scrollbar(self, command=self.__listBox_info.yview)
            self.__Scrollbar3.grid(column=2, row=10, sticky=NSEW)
            self.__listBox_info.configure(yscrollcommand=self.__Scrollbar3.set)

            self.__lb_Alugado = Label(
                self,
                text="Ultimos empréstimos:",
                font=fonttitles,
                bg="White",
                fg=textbgcolor,
            )
            self.__lb_Alugado.grid(column=1, row=12, columnspan=2, sticky=SW)
            self.__tv_emprestimos = ttk.Treeview(
                self, selectmode=BROWSE, columns=(1, 2, 3), show="headings"
            )
            self.__tv_emprestimos.heading(1, text="Desde")
            self.__tv_emprestimos.column(1, stretch=False, width=85, anchor=CENTER)
            self.__tv_emprestimos.heading(2, text="Até")
            self.__tv_emprestimos.column(2, stretch=False, width=85, anchor=CENTER)
            self.__tv_emprestimos.heading(3, text="Nome")
            self.__tv_emprestimos.column(3, stretch=True, width=85)
            self.__tv_emprestimos.grid(row=13, column=1, columnspan=1, sticky=NSEW)
            self.__Scrollbar2 = Scrollbar(self, command=self.__tv_emprestimos.yview)
            self.__Scrollbar2.grid(column=2, row=13, sticky=NSEW)
            self.__tv_emprestimos.configure(yscrollcommand=self.__Scrollbar2.set)

        def Atualizar(self, event):

            self.__tbox_nome.configure(state="normal")
            self.__tbox_nome.delete(0, END)
            self.__tbox_nome.insert(0, self.tabela.Get_Nome())
            self.__tbox_nome.configure(state="readonly")

            self.__tbox_nLivros.configure(state="normal")
            self.__tbox_nLivros.delete(0, END)
            self.__tbox_nLivros.insert(0, self.tabela.Get_NumeroElementos())
            self.__tbox_nLivros.configure(state="readonly")

    class Frame_Editar(Frame):

        """
        Cria uma nova interface para mostrar as informações sobre a tabela aberta;
        """

        def __init__(self, master, tabela):

            super().__init__(master, bg="white")
            pass

    class Janela_nome(Toplevel):
        def __init__(self, Master, frm_bt, notebook):

            super().__init__(master=Master, bg="white")

            self.__master = Master
            self.tab_lst = [notebook.tab(i, option="text") for i in notebook.tabs()]
            self.title("Gestor de Livros")
            self.resizable(width=False, height=False)

            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()

            fonttitles = (
                Settings.Get("Tab_Font_1")["Font"],
                Settings.Get("Tab_Font_1")["T"],
                "italic",
            )
            fonttext = (
                Settings.Get("Tab_Font_2")["Font"],
                Settings.Get("Tab_Font_1")["T"],
            )
            textbgcolor = Settings.Get("Tab_Font_1")["FG"]
            textboxbgcolor = "#d9d9d9"

            self.geometry(
                "300x115+{}+{}".format(
                    (screen_width - 400) // 2, (screen_height - 150) // 2
                )
            )

            Label(
                self,
                background="White",
                text="Indique o nome da área de trabalho:",
                font=fonttitles,
            ).place(x=20, y=10)

            self.Entry = Entry(self, command=None, background=textboxbgcolor)
            self.Entry.place(x=20, y=40, width=260)
            self.Button = Button(
                self,
                text="Criar",
                command=lambda frm_bt=frm_bt, notebook=notebook: self.__Criar(
                    frm_bt, notebook
                ),
            )
            self.Button.place(x=170, y=70, width=80)
            self.bc = Button(self, text="Cancelar", command=self.destroy)
            self.bc.place(x=45, y=70, width=80)

            self.Button.bind(
                "<Enter>",
                lambda event, arg=self.Button: self.__Buttons_Enter(event, arg),
            )
            self.Button.bind(
                "<Leave>",
                lambda event, arg=self.Button: self.__Buttons_Leave(event, arg),
            )

            self.bc.bind(
                "<Enter>", lambda event, arg=self.bc: self.__Buttons_Enter(event, arg)
            )
            self.bc.bind(
                "<Leave>", lambda event, arg=self.bc: self.__Buttons_Leave(event, arg)
            )

            self.configure(background="white")

        def __Buttons_Enter(self, event, button):

            button.config(background=Settings.Get("Buttons_congig")["Selected"])

        def __Buttons_Leave(self, event, button):

            button.config(background=Settings.Get("Buttons_congig")["BG"])

        def __Criar(self, frm_bt, notebook):

            s = self.get_text()
            if s != "" and s != None:
                frm_bt.place_forget()
                notebook.add(MyWidgets.Tab(notebook, s), text=s)
                self.destroy()

        def get_text(self):

            s = self.Entry.get()
            if self.Entry.get().strip() != "":

                count = 0
                for st in self.tab_lst:
                    if st.find(s) != -1:
                        count += 1

                if count == 0:
                    return s
                else:
                    return s + " (" + str(count) + ")"

            else:
                self.attributes("-topmost", False)
                messagebox.showwarning(
                    "Aviso",
                    "Não é posivel adicionar uma area de trabalho sem sem um nome!.",
                )
                self.attributes("-topmost", True)
                return None

    class Janela_Adicionar(Toplevel):

        """
        Abre uma nova janela com a frame adicionar
        """

        def __init__(self, master, tabela):

            super().__init__(master=master, bg="white")

            self.title("Adiconar Livro")

            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()

            self.geometry(
                "300x600+{}+{}".format(
                    (screen_width - 400) // 2, (screen_height - 600) // 2
                )
            )
            self.minsize(300, 500)
            # self.resizable(width=False,height=False)
            self.configure(background="white")

            self.__Frame_adicionar = MyWidgets.Frame_Adicionar(self, tabela)
            self.__Frame_adicionar.pack(fil=BOTH, expand=True)

    class Janela_Frame_FichaTecnica(Toplevel):

        """
        Abre uma nova janela com a frame ficha tecnica
        """

        def __init__(self, master, tabela, livro):

            super().__init__(master=master, bg="white")

            # self.title("Livro")

            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()

            self.geometry(
                "300x600+{}+{}".format(
                    (screen_width - 400) // 2, (screen_height - 600) // 2
                )
            )
            self.minsize(300, 600)
            self.configure(background="white")

            self.__Frame_Ficha = MyWidgets.Frame_FichaTecnica(self, tabela, livro)
            self.__Frame_Ficha.pack(fill=BOTH, expand=True)

    class Janela_Frame_Editar(Toplevel):

        """
        Abre uma nova janela com a frame ficha tecnica
        """

        def __init__(self, master, tabela, livro):

            super().__init__(master=master, bg="white")

    class BarraDePrucura(Frame):
        def __init__(self, master):

            super().__init__(master=master, bg="light gray", height=30)

            self.rowconfigure(0, minsize=30, weight=0)

            self.columnconfigure(0, minsize=100, uniform="esp")
            self.columnconfigure(2, minsize=80, weight=0)
            self.columnconfigure(3, minsize=100, uniform="esp")
            self.columnconfigure(1, minsize=200, weight=1)

            self.__entry = Entry(self, command=None, relief=SUNKEN, bg="gray")
            self.__entry.grid(row=0, column=1, sticky=EW)

            self.__bt_procurar = Button(self, text="Procurar")
            self.__bt_procurar.grid(row=0, column=2, sticky=EW)

    class MenuBarVertical(Frame):
        def __init__(self, master):

            super().__init__(master=master, bg="light gray", width=30)

            self.rowconfigure(0, minsize=30, weight=0)
            self.rowconfigure((2, 4, 6, 8, 10, 12, 14), minsize=30, weight=0)
            self.rowconfigure((1, 3, 5, 7, 9, 11, 13, 15), minsize=2, weight=0)
            self.columnconfigure(0, minsize=30, weight=0)

            Frame(
                self, bg=Settings.Get("Lyt_Clrs_Tebela")["Top_BG"], borderwidth=0
            ).grid(column=0, row=0, sticky=NSEW)

            self.i1 = PhotoImage(file="Program data/imagens/add.png")
            self.B1 = Button(
                self,
                relief=FLAT,
                command=lambda button=1: self.__button_choice(button),
                image=self.i1,
            )
            self.B1.grid(column=0, row=2, sticky=NSEW)
            self.B1.bind("<Enter>", self.__B1_in)
            self.B1.bind("<Leave>", self.__B1_out)

            self.i2 = PhotoImage(file="Program data/imagens/remover.png")
            self.B2 = Button(
                self,
                relief=FLAT,
                state="disable",
                command=lambda button=2: self.__button_choice(button),
                image=self.i2,
            )
            self.B2.grid(column=0, row=4, sticky=NSEW)
            self.B2.bind("<Enter>", self.__B2_in)
            self.B2.bind("<Leave>", self.__B2_out)

            Button(
                self,
                relief=FLAT,
                command=lambda button=3: self.__button_choice(button),
                text="e",
            ).grid(column=0, row=6, sticky=NSEW)
            Button(
                self,
                relief=FLAT,
                command=lambda button=4: self.__button_choice(button),
                text="vr",
            ).grid(column=0, row=8, sticky=NSEW)
            Button(self, relief=FLAT, command=None, text="+s").grid(
                column=0, row=10, sticky=NSEW
            )
            Button(self, relief=FLAT, command=None, text="-s").grid(
                column=0, row=12, sticky=NSEW
            )
            Button(self, relief=FLAT, command=None, text="s").grid(
                column=0, row=14, sticky=NSEW
            )

        def __B1_in(self, event):
            self.i1 = PhotoImage(file="Program data/imagens/add_black.png")
            self.B1.configure(image=self.i1)

        def __B1_out(self, event):
            self.i1 = PhotoImage(file="Program data/imagens/add.png")
            self.B1.configure(image=self.i1)

        def __B2_in(self, event):
            self.i2 = PhotoImage(file="Program data/imagens/remover_black.png")
            self.B2.configure(image=self.i2)

        def __B2_out(self, event):
            self.i2 = PhotoImage(file="Program data/imagens/remover.png")
            self.B2.configure(image=self.i2)

        def __button_choice(self, button):
            self.event_generate("<<Button_Choiche>>", x=button)

    class EditImage(Toplevel):

        Counter = 0

        """
        Abre uma nova janela onde se pode editar uma imagem 
        """

        def __init__(self, master, image_path, func):

            MyWidgets.EditImage.Counter += 1

            super().__init__(master)

            if not isinstance(image_path, str):
                raise ValueError("image_path not diretory")

            self.title("Editar Imagem")

            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()

            self.geometry(
                "500x500+{}+{}".format(
                    (screen_width - 400) // 2, (screen_height - 600) // 2
                )
            )
            self.minsize(500, 500)
            self.resizable(width=False, height=False)
            self.configure(background="white")
            self.attributes("-topmost", True)
            self.attributes("-topmost", False)

            # botoes guardar e cancelar
            Button(
                self, command=lambda func=func: self.concluir(func), text="Concluir"
            ).place(x=260, y=465, width=100, height=25)
            Button(self, command=self.cancelar, text="Cancelar").place(
                x=150, y=465, width=100, height=25
            )

            # canvas
            self.canvas = Canvas(
                self, highlightthickness=2, highlightbackground="black"
            )
            self.canvas.place(x=10, y=5, width=480, height=450)

            self.image_path = image_path
            self.image_ori = Image.open(image_path)

            if self.image_ori.width < 300 or self.image_ori.height < 270:
                self.image_temp = self.image_ori.resize((300, 270))
                self.image_ori = self.image_temp
                self.image_in = self.image_ori
                self.img = ImageTk.PhotoImage(self.image_temp)
            else:
                self.image_in = self.image_ori
                self.img = ImageTk.PhotoImage(self.image_ori)

            self.image_resize = self.image_in

            self.Id = self.canvas.create_image(240, 225, anchor=CENTER, image=self.img)
            self.canvas.tag_bind(self.Id, "<B1-Motion>", self.ON_MOVE_PRESS)
            self.canvas.tag_bind(self.Id, "<Button-1>", self.ON_PRESS)

            self.canvas.create_rectangle(
                90, 90, 390, 360, fill=None, outline="red", width=2
            )

            # botões Mover, Resize e Crop
            self.b_Move = Button(self, command=self.Mover, text="Mover")
            self.b_Move.place(x=12, y=428, width=50, height=25)
            self.B_b_Move = False

            self.b_Resize = Button(self, command=self.Resize, text="Resize")
            self.b_Resize.place(x=64, y=428, width=50, height=25)
            self.B_b_Resize = False

            self.b_crop = Button(self, command=self.Crop, text="Crop")
            self.b_crop.place(x=116, y=428, width=50, height=25)

            self.b_reset = Button(self, command=self.Reset, text="Reset")
            self.b_reset.place(x=168, y=428, width=50, height=25)

            self.MousePos = (0, 0)

        def ON_PRESS(self, event):

            self.MousePos = (event.x, event.y)

        def ON_MOVE_PRESS(self, event):

            if self.B_b_Move:

                pos = self.canvas.coords(self.Id)

                aresta_cima = (
                    pos[1] + event.y - self.MousePos[1]
                ) - self.image_in.height / 2
                aresta_baixo = (
                    pos[1] + event.y - self.MousePos[1]
                ) + self.image_in.height / 2
                aresta_esquerda = (
                    pos[0] + event.x - self.MousePos[0]
                ) - self.image_in.width / 2
                aresta_direita = (
                    pos[0] + event.x - self.MousePos[0]
                ) + self.image_in.width / 2

                # centro 240, 225
                # E90,D390,C90,B360
                # CMB EMD
                # Meio direita
                # canto superior direito
                if (aresta_cima >= 90) and (aresta_direita <= 390):
                    self.canvas.coords(
                        self.Id,
                        (390 - self.image_in.width / 2, 90 + self.image_in.height / 2),
                    )
                # canto superior esquerdo
                elif (aresta_cima >= 90) and (aresta_esquerda >= 90):
                    self.canvas.coords(
                        self.Id,
                        (90 + self.image_in.width / 2, 90 + self.image_in.height / 2),
                    )
                # canto inferior direito
                elif (aresta_baixo <= 360) and (aresta_direita <= 390):
                    self.canvas.coords(
                        self.Id,
                        (390 - self.image_in.width / 2, 360 - self.image_in.height / 2),
                    )
                # canto inferior esquerdo
                elif (aresta_baixo <= 360) and (aresta_esquerda >= 90):
                    self.canvas.coords(
                        self.Id,
                        (90 + self.image_in.width / 2, 360 - self.image_in.height / 2),
                    )
                # Meio Direita
                elif aresta_direita <= 390:
                    self.canvas.coords(
                        self.Id,
                        (
                            390 - self.image_in.width / 2,
                            pos[1] + event.y - self.MousePos[1],
                        ),
                    )
                # Meio esquerda
                elif aresta_esquerda >= 90:
                    self.canvas.coords(
                        self.Id,
                        (
                            90 + self.image_in.width / 2,
                            pos[1] + event.y - self.MousePos[1],
                        ),
                    )
                # Cima Meio
                elif aresta_cima >= 90:  # self.image_in.height/2 + 450 :
                    self.canvas.coords(
                        self.Id,
                        (
                            pos[0] + event.x - self.MousePos[0],
                            90 + self.image_in.height / 2,
                        ),
                    )
                # Baixo Meio
                elif aresta_baixo <= 360:
                    self.canvas.coords(
                        self.Id,
                        (
                            pos[0] + event.x - self.MousePos[0],
                            360 - self.image_in.height / 2,
                        ),
                    )
                # Meio Meio
                else:
                    self.canvas.coords(
                        self.Id,
                        (
                            pos[0] + event.x - self.MousePos[0],
                            pos[1] + event.y - self.MousePos[1],
                        ),
                    )

                self.MousePos = (event.x, event.y)

            elif self.B_b_Resize:

                pos = self.canvas.coords(self.Id)

                aresta_cima = (
                    pos[1] + event.y - self.MousePos[1]
                ) - self.image_in.height / 2
                aresta_baixo = (
                    pos[1] + event.y - self.MousePos[1]
                ) + self.image_in.height / 2
                aresta_esquerda = pos[0] - self.image_in.width / 2
                aresta_direita = pos[0] + self.image_in.width / 2

                delta_y = event.y - self.MousePos[1]
                delta_x = event.x - self.MousePos[0]
                w = self.image_in.width
                h = self.image_in.height

                if w + delta_x <= 300:
                    new_w = 300
                else:
                    new_w = w + delta_x

                if h + delta_y <= 270:
                    new_h = 270
                else:
                    new_h = h + delta_y

                self.image_temp = self.image_resize.resize((new_w, new_h))

                # centro 240, 225
                # E90,D390,C90,B360
                # CMB EMD
                # Meio direita
                # canto superior direito
                if (aresta_cima >= 90) and (aresta_direita <= 390):
                    self.canvas.coords(
                        self.Id,
                        (
                            390 - self.image_temp.width / 2,
                            90 + self.image_temp.height / 2,
                        ),
                    )
                # canto superior esquerdo
                elif (aresta_cima >= 90) and (aresta_esquerda >= 90):
                    self.canvas.coords(
                        self.Id,
                        (
                            90 + self.image_temp.width / 2,
                            90 + self.image_temp.height / 2,
                        ),
                    )
                # canto inferior direito
                elif (aresta_baixo <= 360) and (aresta_direita <= 390):
                    self.canvas.coords(
                        self.Id,
                        (
                            390 - self.image_temp.width / 2,
                            360 - self.image_temp.height / 2,
                        ),
                    )
                # canto inferior esquerdo
                elif (aresta_baixo <= 360) and (aresta_esquerda >= 90):
                    self.canvas.coords(
                        self.Id,
                        (
                            90 + self.image_temp.width / 2,
                            360 - self.image_temp.height / 2,
                        ),
                    )
                # Meio Direita
                elif aresta_direita <= 390:
                    self.canvas.coords(
                        self.Id,
                        (
                            390 - self.image_temp.width / 2,
                            pos[1] + event.y - self.MousePos[1],
                        ),
                    )
                # Meio esquerda
                elif aresta_esquerda >= 90:
                    self.canvas.coords(
                        self.Id,
                        (
                            90 + self.image_temp.width / 2,
                            pos[1] + event.y - self.MousePos[1],
                        ),
                    )
                # Cima Meio
                elif aresta_cima >= 90:  # self.image_in.height/2 + 450 :
                    self.canvas.coords(
                        self.Id,
                        (
                            pos[0] + event.x - self.MousePos[0],
                            90 + self.image_in.height / 2,
                        ),
                    )
                # Baixo Meio
                elif aresta_baixo <= 360:
                    self.canvas.coords(
                        self.Id,
                        (
                            pos[0] + event.x - self.MousePos[0],
                            360 - self.image_in.height / 2,
                        ),
                    )
                # Meio Meio

                self.image_in = self.image_temp
                self.img = ImageTk.PhotoImage(self.image_temp)
                self.canvas.itemconfig(self.Id, image=self.img)

                self.MousePos = (event.x, event.y)

        def Reset(self):

            self.image_resize = self.image_ori
            self.image_in = self.image_ori
            self.img = ImageTk.PhotoImage(self.image_ori)
            self.canvas.itemconfig(self.Id, image=self.img)
            self.canvas.coords(self.Id, (240, 225))

        def Mover(self):

            self.B_b_Move = not self.B_b_Move

            if self.B_b_Move:

                self.b_Move.configure(relief=SUNKEN, background="light grey")

                self.B_b_Resize = False
                self.b_Resize.configure(relief=RAISED, background="#f0f0f0")

            else:

                self.b_Move.configure(relief=RAISED, background="#f0f0f0")

        def Resize(self):

            self.B_b_Resize = not self.B_b_Resize

            if self.B_b_Resize:

                self.b_Resize.configure(relief=SUNKEN, background="light grey")

                self.B_b_Move = False
                self.b_Move.configure(relief=RAISED, background="#f0f0f0")

            else:

                self.b_Resize.configure(relief=RAISED, background="#f0f0f0")

        def Crop(self):

            self.B_b_crop = True

            self.B_b_Move = False
            self.b_Move.configure(relief=RAISED, background="#f0f0f0")

            self.B_b_crop = False
            self.b_crop.configure(relief=RAISED, background="#f0f0f0")

            pos = self.canvas.coords(self.Id)
            left_x = int(pos[0] - self.image_in.width / 2)
            left_y = int(pos[1] - self.image_in.height / 2)

            if left_x >= 0:
                x = 90 - left_x
            else:
                x = -left_x + 90

            if left_y >= 0:
                y = 90 - left_y
            else:
                y = -left_y + 90

            self.image_temp = self.image_in.crop((x, y, x + 300, y + 270))
            self.image_in = self.image_temp
            self.img = ImageTk.PhotoImage(self.image_temp)
            self.canvas.itemconfig(self.Id, image=self.img)
            self.canvas.coords(self.Id, (240, 225))

            self.image_resize = self.image_in

        def concluir(self, func):
            func(self.image_in)

        def cancelar(self):
            MyWidgets.EditImage.Counter -= 1
            self.destroy()


class Livro:

    __Estados = {"1": "Disponivel", "0": "indisponivel"}
    __Conditions = {}

    def __init__(
        self,
        numero,
        title,
        autor="n/d",
        editora="n/d",
        genero="n/d",
        edition="n/d",
        datain="n/d",
        dataout="n/d",
        condition="Bom Estado",
        imagem="n/d",
        outros="n/d",
        hist=Stack(),
        histp=Stack(),
    ):

        self.__Title = title
        self.__Autor = autor
        self.__Editora = editora
        self.__Genero = genero
        self.__Edition = edition
        self.__DataIn = datain
        self.__DataOut = dataout
        self.__Condition = condition
        self.__outros = outros

        self.__numero = numero

        self.__Historico = hist
        self.__HistoricoEmprestimos = histp

        self.__Estado = self.__Estados["1"]

        if isinstance(imagem, str):
            self.__Imagem = imagem
        else:
            self.__Imagem = None

        Manager.WInfo("Livro Criado", WInfoDraw=None, linha=500, Prioridade=2)

    def Novo_Emprestimo(self, Nome, data, condition=None):

        if condition == None:
            condition = self.__Condition
        else:
            self.__Condition = condition

        if not isinstance(Nome, str):
            Manager.WInfo(
                "Empréstmo, Nome Inválido. Não efetuado",
                WInfoDraw=None,
                linha=504,
                Prioridade=1,
            )
            return False
        else:
            if isinstance(data, tuple):
                if len(data) == 3:
                    time = datetime.now()
                    dic = {
                        "Nome": nome,
                        "DataOut": time.strftime("%d/%m/%Y"),
                        "Hora": time.strftime("%H:%M:%S"),
                        "DataLim": "{0}/{1}/{2}".format(data[0], data[1], data[2]),
                        "Condição": self.__Condition,
                    }
                    self.__HistoricoEmprestimos.Adicionar(dic)
                    self.__Historico.Adicionar(
                        {
                            "Data": dic["Data"],
                            "Hora": dic["Hora"],
                            "Evento": "Empréstimo",
                            "TxT": None,
                        }
                    )
                    del time
                    del dic
                    self.__Estado = self.__Estados["0"]
                    Manager.WInfo(
                        "Empréstimo, Concluído", WInfoDraw=None, linha=517, Prioridade=1
                    )
                    return True
                else:
                    Manager.WInfo(
                        "Empréstimo, Data Inválida",
                        WInfoDraw=None,
                        linha=517,
                        Prioridade=1,
                    )
                    return False
            else:
                Manager.WInfo(
                    "Empréstimo, Data Inválida", WInfoDraw=None, linha=521, Prioridade=1
                )
                return False

    def Cancelar_Emepréstimo(self):
        self.__HistoricoEmprestimos.Remover()
        self.__Historico.Remover()
        self.__Estado = self.__Estados["1"]

    def Reset_Emprestimos(self):
        self.__HistoricoEmprestimos = Stack()

    def Novo_Historial(self, info, evento="Histórico"):

        if isinstance(info, str):
            time = datetime.now()
            self.__Historico.Adicionar(
                {
                    "Data": dic["Data"],
                    "Hora": dic["Hora"],
                    "Evento": evento,
                    "TxT": info,
                }
            )
            del time
            Manager.WInfo(
                "Informações do livro, atualizadas.",
                WInfoDraw=None,
                linha=537,
                Prioridade=2,
            )
            return True
        else:
            Manager.WInfo(
                "Informações do livro, NÃO FORAM atualizadas.",
                WInfoDraw=None,
                linha=540,
                Prioridade=1,
            )
            return False

    def Retirar_Historial(self):
        self.__Historico.Remover()

    def Reset_Historial(self):
        self.__Historico = Stack()

    # Geters
    def Get_Title(self):
        return self.__Title

    def Get_Autor(self):
        return self.__Autor

    def Get_Genero(self):
        return self.__Genero

    def Get_Edition(self):
        return self.__Edition

    def Get_Editora(self):
        return self.__Editora

    def Get_DataIn(self):
        return self.__DataIn

    def Get_DataOut(self):
        return self.__DataOut

    def Get_Historico(self):
        return self.__Historico

    def Get_Emprestimos_Todos(self):
        return self.__HistoricoEmprestimos

    def Get_Emprestimos_Numero(self):
        return self.__HistoricoEmprestimos.Tamanho()

    def Get_Condition(self):
        return self.__Condition

    def Get_Estado(self):
        return self.__Estado

    def Get_Numero(self):
        return self.__numero

    def Get_ImagemPath(self):
        return self.__Imagem

    def Get_Outros(self):
        return self.__outros

    def Get_Basic_Info(self):
        return {
            "Title": self.__Title,
            "Autor": self.__Autor,
            "Editora": self.__Editora,
            "Genero": self.__Genero,
            "Edition": self.__Edition,
            "Data de Entrada": self.__DataIn,
            "Data da Edição": self.__DataOut,
            "Outros": self.__outros,
        }

    # Setters
    def Set_Title(self, name):
        if isinstance(name, str):
            self.__Title = name

    def Set_Autor(self, name):
        if isinstance(name, str):
            self.__Autor = name

    def Set_Genero(self, name):
        if isinstance(name, str):
            self.__Genero = name

    def Set_Edition(self, n):
        if isinstance(n, int):
            self.__Edition = n

    def Set_Editora(self, name):
        if isinstance(name, str):
            self.__Editora = name

    def Set_DataIn(self, dd, mm, aa):
        if isinstance(dd, int) and isinstance(mm, int) and isinstance(aa, int):
            self.__DataIn = "{0}/{1}/{2}".format(dd, mm, aa)

    def Set_DataOut(self, dd, mm, aa):
        if isinstance(dd, int) and isinstance(mm, int) and isinstance(aa, int):
            self.__DataOut = "{0}/{1}/{2}".format(dd, mm, aa)

    def Set_Estado(self, estado):
        if estado in self.__Conditions:
            self.__Condition = condition

    def Set_Outros(self, name):
        if isinstance(name, str):
            self.__outros = name

    def Set_ImagemPath(self, imagem):
        self.__Imagem = imagem


class Manager:
    class Crash_Handler:
        # janela chandler e metodos e atributos
        pass

    class WInfoDraw:

        __Historico = ()

        def __init__(self, Master, posx, posy, bg=None, fg=None, animado=False):

            self.__Stack = Stack()  # nao faz sentido
            self.__animado = animado

            self.__Frm = Frame(Barra, width=400, height=20, bg="red")
            self.__lb_WinfoText = Label(
                Frm,
                text="Este é o texto de teste da class não tem mais que escreva",
                wraplength=380,
                justify="center",
                background="red",
            )
            self.__lb_WinfoText2 = Label(
                Frm, text="", wraplength=380, justify="center", background="red"
            )
            self.__C_WinfoIcon = Canvas(
                Frm, width=20, height=20, bg=bg, highlightthickness=0
            )
            self.__C_WinfoIcon2 = Canvas(
                Frm, width=20, height=20, bg=bg, highlightthickness=0
            )

            self.__lb_WinfoText.place(x=20, y=0)
            self.__C_WinfoIcon.place(x=0, y=0)

            Frm.place(relx=pos, y=posy)

        def New_WInfo(self, winfo):

            if self.__animado:
                pass

        def Escrever_AnimadoUP(self):

            for i in range(18, -2, -2):

                self.__lb_WinfoText.place_forget()
                self.__lb_WinfoText.place(x=0, y=i - 20)
                self.__C_WinfoIcon.place_forget()
                self.__C_WinfoIcon.place(x=0, y=i - 20)

                self.__lb_WinfoText2.place_forget()  # mostra o primeiro do stack
                self.__lb_WinfoText2.config(
                    text=self.__Stack.Primeiro().get_Text(),
                    foreground=self.__Stack.Primeiro().get_WinfoColor(),
                )
                self.__lb_WinfoText2.place(x=0, y=i)
                self.__C_WinfoIcon.place_forget()
                # self.__C_WinfoIcon2.config()
                self.__C_WinfoIcon.place(x=0, y=i)

                self.__lb_WinfoText.update_idletasks()
                self.__lb_WinfoText2.update_idletasks()
                self.__C_WinfoIcon.update_idletasks()
                self.__C_WinfoIcon2.update_idletasks()

                # Manager.WInfo.get_WinfoColor

            troca = self.__lb_WinfoText
            self.__lb_WinfoText = self.__lb_WinfoText2
            self.__lb_WinfoText2 = troca

            troca = self.__C_WinfoIcon
            self.__C_WinfoIcon = self.__C_WinfoIcon2
            self.__C_WinfoIcon2 = troca

        def Escrever():

            self.__lb_WinfoText.config(
                text=self.__Stack.Primeiro().get_Text(),
                foreground=self.__Stack.Primeiro().get_WinfoColor(),
            )
            # self.__C_WinfoIcon2.config()

    class WInfo:
        def __init__(self, Text, WInfoDraw, linha, Prioridade=1):

            """
            prioridades: (3-info1,2-info2,1-warnig,0-Alerta Critico)
            """

            self.__Prioridade = Prioridade
            self.__Text = Text
            self.__WinfoIcon = Settings.Get("Path_Winfo_Images")[
                Prioridade
            ]  # mudar para pfotoimage
            self.__WinfoColor = Settings.Get("Path_Winfo_Images")[Prioridade]
            time = datetime.now()
            self.__Time = time.strftime("%H:%M:%S")
            self.__Date = time.strftime("%d/%m/%Y")
            del time
            self.__Linha = linha
            # self.Adicionar_WInfo(self)
            # self.__Stack.Adicionar(self)
            # self.Escrever_Novo()

        def get_Linha(self):
            return self.__Linha

        def get_Text(self):
            return self.__Text

        def get_Prioridade(self):
            return self.__Prioridade

        def get_WinfoIcon(self):
            return self.__WinfoIcon

        def get_WinfoColor(self):
            return self.__WinfoColor

        def get_Time(self):
            return self.__Time

        def get_Date(self):
            return self.__Date

        def Change_Text(self, text):
            if isinstance(text, str):
                self.__Text = text

        def Change_Prioridade(self, Prioridade):
            if prioridade in (0, 1, 2, 3):
                self.__Prioridade = Prioridade
                self.__WinfoIcon = Settings.Get("Path_Winfo_Images")[Prioridade]
                self.__WinfoColor = Settings.Get("Path_Winfo_Images")[Prioridade]


class Jan_Main:
    def __init__(self):

        Janela_principal = Tk()
        Janela_principal.title("Gestor De Livros")

        screen_width = Janela_principal.winfo_screenwidth()
        screen_height = Janela_principal.winfo_screenheight()

        Janela_principal.resizable(width=True, height=True)
        Janela_principal.minsize(800, 650)
        Janela_principal.maxsize(screen_width, screen_height)
        Janela_principal.state("zoomed")
        Janela_principal.geometry("800x600+0+0")

        Janela_principal.columnconfigure(0, minsize=800, weight=1)
        Janela_principal.rowconfigure(0, minsize=100, weight=1)
        Janela_principal.rowconfigure(1, minsize=24, weight=0)

        # criando o menu
        self.__CriarMenu(Janela_principal)
        Janela_principal.config(menu=self.menubar)  # aplicando

        # Criando Barra de baixo
        self.__CriarBarra(Janela_principal)

        # Criando Frame Base
        self.tablist = []
        self.__CriarTabFrame(Janela_principal)

        #   Janela_principal.bind("<Configure>", lambda event, Master=Janela_principal : self.OnResize(event, Master=Master))

        style = ttk.Style()
        style.theme_use("default")
        style.map("Treeview")
        style.map("ComboBox")

        # Criando butões e texto
        if Settings.LastSession() != 1:
            self.__Buttons(Janela_principal)

        else:
            pass

        Janela_principal.mainloop()

    def __CriarMenu(self, top):

        # Barra Menu NÃO É CONFIGURAVEL
        self.menubar = Menu(top, tearoff=0)

        # Menu Ficheiros accelerator="Ctrl+Q") image
        self.filemenu = Menu(self.menubar, tearoff=0)
        self.filemenu.config(
            foreground=Settings.Get("Font_Menu")["FG"],
            background=Settings.Get("Font_Menu")["BG"],
            font=(Settings.Get("Font_Menu")["Font"], Settings.Get("Font_Menu")["T"]),
        )
        self.filemenu.add_command(label="Novo", command=self.Nada)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Abrir", command=self.Nada)
        self.filemenu.add_command(label="Abrir Recente", command=self.Nada)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Guardar", command=self.Nada)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Fechar", command=self.Nada)
        # Adicionando ao base
        self.menubar.add_cascade(label="Ficheiro", menu=self.filemenu)

        # Menu Preferências
        self.preferenciasmenu = Menu(self.menubar, tearoff=0)
        self.preferenciasmenu.config(
            foreground=Settings.Get("Font_Menu")["FG"],
            background=Settings.Get("Font_Menu")["BG"],
            font=(Settings.Get("Font_Menu")["Font"], Settings.Get("Font_Menu")["T"]),
        )
        self.preferenciasmenu.add_command(
            label="Editar Preferências", command=self.Nada
        )
        # Adicionando ao base
        self.menubar.add_cascade(label="Editar", menu=self.preferenciasmenu)

        # Menu lista
        self.listamenu = Menu(self.menubar, tearoff=0)
        self.listamenu.config(
            foreground=Settings.Get("Font_Menu")["FG"],
            background=Settings.Get("Font_Menu")["BG"],
            font=(Settings.Get("Font_Menu")["Font"], Settings.Get("Font_Menu")["T"]),
        )
        self.listamenu.add_command(
            label="Nova", command=lambda Master=top: self.NovaTab(Master)
        )
        self.listamenu.add_command(label="Abrir", command=self.AbrirTab)
        self.listamenu.add_separator()
        self.listamenu.add_command(label="Editar", command=self.Nada)
        self.listamenu.add_separator()
        self.listamenu.add_command(label="Apagar", command=self.Nada)
        self.listamenu.add_separator()
        self.listamenu.add_command(label="Guardar", command=self.GuardarTab)
        self.listamenu.add_command(label="Guardar em", command=self.GuardarEmTab)
        self.listamenu.add_separator()
        self.listamenu.add_command(label="Fechar ", command=self.Nada)
        # Adicionando ao base
        self.menubar.add_cascade(label="Lista", menu=self.listamenu)

        # Menu Livro
        self.livromenu = Menu(self.menubar, tearoff=0)
        self.livromenu.config(
            foreground=Settings.Get("Font_Menu")["FG"],
            background=Settings.Get("Font_Menu")["BG"],
            font=(Settings.Get("Font_Menu")["Font"], Settings.Get("Font_Menu")["T"]),
        )
        self.livromenu.add_command(label="Adicionar", command=self.Nada)
        self.livromenu.add_separator()
        self.livromenu.add_command(label="Procurar", command=self.Nada)
        self.livromenu.add_separator()
        self.livromenu.add_command(label="Remover da Lista")
        self.livromenu.add_command(label="Remover de Todas")
        # Adicionando ao base
        self.menubar.add_cascade(label="Livro", menu=self.livromenu)

        # butao ajuda
        self.menubar.add_command(label="Ajuda", command=self.Nada)
        # inutil não altera  menubar.configure(takefocus="Ajudar",foreground="red"

    def __CriarBarra(self, Master):

        self.Barra = Frame(Master, height=24, bg="white")
        lb_Verção = Label(self.Barra, text="Versão: " + Settings.Versao(), bg="white")
        lb_Verção.place(x=10, rely=0.125)
        Master.update_idletasks()
        self.Barra.grid(row=1, column=0, sticky=NSEW)

    def __CriarTabFrame(self, Master):

        Master.update_idletasks()
        self.notebook = ttk.Notebook(master=Master)
        self.notebook.grid(row=0, column=0, sticky=NSEW)

    def __Buttons(self, Master):

        self.frm_bt = Frame(Master, width=300, height=80, background="#D9D9D9")

        self.B_Abrir_Area = Button(
            self.frm_bt,
            text="Abrir",
            command=self.AbrirTab,
            background=Settings.Get("Buttons_congig")["BG"],
        )
        self.B_Abrir_Area.config(
            foreground=Settings.Get("Font_Buttons")["FG"],
            background=Settings.Get("Font_Buttons")["BG"],
            font=(
                Settings.Get("Font_Buttons")["Font"],
                Settings.Get("Font_Buttons")["T"],
            ),
        )
        self.B_Abrir_Area.place(x=33, y=50, width=100, height=25)
        self.B_Abrir_Area.bind(
            "<Enter>",
            lambda event, arg=self.B_Abrir_Area: self.__Buttons_Enter(event, arg),
        )
        self.B_Abrir_Area.bind(
            "<Leave>",
            lambda event, arg=self.B_Abrir_Area: self.__Buttons_Leave(event, arg),
        )

        self.B_Nova_Area = Button(
            self.frm_bt,
            text="Nova Área",
            command=lambda Master=Master: self.NovaTab(Master),
        )
        self.B_Nova_Area.config(
            foreground=Settings.Get("Font_Buttons")["FG"],
            background=Settings.Get("Font_Buttons")["BG"],
            font=(
                Settings.Get("Font_Buttons")["Font"],
                Settings.Get("Font_Buttons")["T"],
            ),
        )
        self.B_Nova_Area.place(x=166, y=50, width=100, height=25)
        self.B_Nova_Area.bind(
            "<Enter>",
            lambda event, arg=self.B_Nova_Area: self.__Buttons_Enter(event, arg),
        )
        self.B_Nova_Area.bind(
            "<Leave>",
            lambda event, arg=self.B_Nova_Area: self.__Buttons_Leave(event, arg),
        )

        self.lb = Label(
            self.frm_bt,
            text="Nenhuma Área De Trabalho Aberta",
            font=(Settings.Get("Font_Menu")["Font"], Settings.Get("Font_Menu")["T"]),
            foreground=Settings.Get("Font_Menu")["FG"],
            background="#D9D9D9",
        )

        self.lb.place(x=0, rely=0, relwidth=1)
        self.frm_bt.place(relx=0.5, rely=0.5, anchor="center")

    def NovaTab(self, Master):

        MyWidgets.Janela_nome(Master, self.frm_bt, notebook=self.notebook)
        self.frm_bt.place_forget()

    def AbrirTab(self):

        path = filedialog.askopenfile(filetypes=[("Excel files", "*.xlsx;")])

        if path != None:

            myWorkbook = load_workbook(path.name, data_only=True)

            c = 0
            for sheet_name in ("Livros", "Image Paths", "Tabela info"):

                if sheet_name in myWorkbook.sheetnames:
                    c += 1

            if c == 3:

                self.frm_bt.place_forget()
                self.frm_bt.forget()

                Wsheet1 = myWorkbook["Livros"]
                Wsheet2 = myWorkbook["Image Paths"]
                Wsheet3 = myWorkbook["Tabela info"]

                mytab = MyWidgets.Tab(self.notebook, Wsheet3["A1"].value, path.name)
                self.notebook.add(mytab, text=Wsheet3["A1"].value)

                c = 2
                while Wsheet1["A" + str(c + 1)].value != None:

                    c += 1
                    numb = str(c)
                    if (
                        Wsheet2["B" + numb].value != "n/d"
                        and Wsheet2["B" + numb].value != None
                    ):

                        p = path.name[: -len(path.name.split("/")[-1])]

                        COPYFILE(
                            p + Wsheet2["B" + numb].value, Wsheet2["B" + numb].value
                        )
                        caminho = Wsheet2["B" + numb].value
                    else:
                        caminho = "n/d"

                    mytab.tabela.Novo_Elemento(
                        Livro(
                            Wsheet1["k" + numb].value,
                            Wsheet1["B" + numb].value,
                            Wsheet1["C" + numb].value,
                            Wsheet1["D" + numb].value,
                            Wsheet1["E" + numb].value,
                            Wsheet1["F" + numb].value,
                            Wsheet1["G" + numb].value,
                            Wsheet1["H" + numb].value,
                            Wsheet1["i" + numb].value,
                            caminho,
                            Wsheet1["J" + numb].value,
                        ),
                        Wsheet1["A" + numb].value,
                    )

                    mytab.frm_InformaçõesTabela.Atualizar(None)

                myWorkbook.close()

            else:
                messagebox.showerror(
                    "Gestor de Livros", "O ficheiro carregado não é compativél"
                )

    def EditarTab(self):

        if len(self.notebook.tabs()) > 0:

            pass

        else:
            messagebox.showwarning(
                "Gestor de Livros",
                "Não pode Editar um Lista sem pelo menos uma aberta !",
            )

    def ApagarTab(self):

        if len(self.notebook.tabs()) > 0:
            pass
        else:
            messagebox.showwarning(
                "Gestor de Livros",
                "Não pode Apagar um Lista sem pelo menos uma aberta !",
            )

    def GuardarTab(self):

        if len(self.notebook.tabs()) > 0:

            tab_id = self.notebook.select()

            if (
                messagebox.askquestion(
                    "Gestor de Livros",
                    "Deseja Guardar a lista: "
                    + self.notebook.tab(tab_id, option="text")
                    + " ?",
                )
                == "yes"
            ):

                self.notebook.event_generate("<<Tab_Save>>", x=1)

    def GuardarEmTab(self):

        if len(self.notebook.tabs()) > 0:

            tab_id = self.notebook.select()

            if (
                messagebox.askquestion(
                    "Gestor de Livros",
                    "Deseja Guardar a lista: "
                    + self.notebook.tab(tab_id, option="text")
                    + " ?",
                )
                == "yes"
            ):

                Path = filedialog.asksaveasfilename()

                if isinstance(Path, str) and Path != "" and not path.exists(Path):

                    NOVODIRETORIO(Path)
                    NOVODIRETORIO(Path + "/Livros.imagens")

                    MyWidgets.Tab.Path = (
                        Path + "?" + self.notebook.tab(tab_id, option="text")
                    )
                    self.notebook.event_generate("<<Tab_Save>>", x=2)

                else:
                    messagebox.showwarning(
                        "Gestor de Livros", Path + "/nNão é um diretório válido !"
                    )
        else:
            messagebox.showwarning(
                "Gestor de Livros",
                "Não pode guardar um Lista sem pelo menos uma aberta !",
            )

    def Fechar(self):
        pass

    def __Buttons_Enter(self, event, button):

        button.config(background=Settings.Get("Buttons_congig")["Selected"])

    def __Buttons_Leave(self, event, button):

        button.config(background=Settings.Get("Buttons_congig")["BG"])

    """
    def OnResize(self,event,Master):
        try:
            self.Barra.update_idletasks()
            self.Barra.place(x=0,rely=(Master.winfo_height()-24)/Master.winfo_height(),relwidth=1,height=24)
            self.notebook.update_idletasks()
            self.notebook.place(x=0,y=0,relheight=(Master.winfo_height()-24)/Master.winfo_height(),relwidth=1)
        except:
            print("Stupid BUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUUG")
    """

    def Nada(self):
        pass

    def __exit__():
        pass


f = Jan_Main()
print("end")
